# procesador_excel.py
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import os
import PyPDF2
from tkinter import StringVar, Canvas
import re
from PIL import Image, ImageTk
import pytesseract
import platform
import cv2
import numpy as np
import pandas as pd
from PyPDF2 import PdfMerger
from openpyxl import load_workbook
import openpyxl
import glob
import fitz  # Para manejar el archivo PDF

# Set TESSDATA_PREFIX environment variable
os.environ['TESSDATA_PREFIX'] = '/usr/local/share/tessdata/'

def load_excel_data_from_sheet(filename, sheet_name, max_rows=50):
    """
    Carga los datos de una hoja específica del archivo Excel, incluyendo valores calculados.
    Solo permite archivos .xlsx
    """
    extension = os.path.splitext(filename)[1].lower()
    data = []

    if extension != '.xlsx':
        raise ValueError("Solo se permiten archivos con extensión .xlsx")
        
    try:
        wb = openpyxl.load_workbook(filename, data_only=True)
        sheet = wb[sheet_name]
        for row in sheet.iter_rows(min_row=1, max_row=max_rows, values_only=True):
            data.append(row)
    except Exception as e:
        print(f"Error al cargar los datos de la hoja: {str(e)}")
    
    return data
def find_header_row(data, required_columns):
    """
    Encuentra la fila de encabezados basada en las columnas requeridas.
    Ahora es más flexible en la búsqueda y muestra más información de depuración.
    """
    print("\nBuscando encabezados...")
    print(f"Columnas requeridas: {required_columns}")
    
    for idx, row in enumerate(data):
        if row is None:
            continue
            
        print(f"\nAnalizando fila {idx + 1}:")
        print(f"Contenido de la fila: {row}")
        
        # Crear una lista de valores limpiados para comparación
        row_values = [str(cell).upper().strip() if cell is not None else '' for cell in row]
        print(f"Valores limpiados: {row_values}")
        
        # Buscar coincidencias para cada columna requerida
        matches = []
        for required_column in required_columns:
            required_column_upper = required_column.upper()
            found = False
            for cell_value in row_values:
                # Verificar coincidencia exacta o si la columna requerida está contenida en el valor
                if cell_value and (required_column_upper == cell_value or required_column_upper in cell_value):
                    matches.append(required_column)
                    print(f"Encontrada columna: {required_column}")
                    found = True
                    break
            
            if not found:
                print(f"No se encontró la columna: {required_column}")
        
        # Si encontramos todas las columnas requeridas
        if len(matches) == len(required_columns):
            print(f"\n¡Encabezados encontrados en la fila {idx + 1}!")
            return idx, row
    
    print("\nNo se encontraron todos los encabezados requeridos.")
    return None, None

def obtener_valor_inventario(file_path, referencia):
        """Obtiene el valor de la referencia en la hoja INVENTARIO."""
        extension = os.path.splitext(file_path)[1].lower()
        
        if extension == ".xlsx":
            wb = openpyxl.load_workbook(file_path, data_only=True)
            inventario = wb["INVENTARIO"]
            
            columna = ''.join(filter(str.isalpha, referencia))
            fila = ''.join(filter(str.isdigit, referencia))
            
            return inventario[f"{columna}{fila}"].value
        else:
            raise ValueError("Solo se permiten archivos con extensión .xlsx")

class PDFProcessorApp:
    def __init__(self, root):
        self.root = root
        self.setup_window()
        self.create_widgets()
        self.ocr_lang = 'eng'
        self.excel_path = None
        self.excel_data = None
        self.clientes_info = {}
        
    def setup_window(self):
        self.root.title("Procesador de PDF y Excel")
        self.root.geometry("600x400")
        self.center_window()
        
    def center_window(self):
        self.root.update_idletasks()
        width = self.root.winfo_width()
        height = self.root.winfo_height()
        x = (self.root.winfo_screenwidth() // 2) - (width // 2)
        y = (self.root.winfo_screenheight() // 2) - (height // 2)
        self.root.geometry(f'{width}x{height}+{x}+{y}')

    def create_widgets(self):
        main_frame = ttk.Frame(self.root, padding="20")
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # Título
        title_label = ttk.Label(
            main_frame,
            text="Procesador de PDF - Subpartidas",
            font=('Helvetica', 16, 'bold')
        )
        title_label.pack(pady=20)
        
        # Excel file input
        self.excel_path_var = StringVar()
        excel_label = ttk.Label(main_frame, text="Archivo Excel:", font=('Helvetica', 10))
        excel_label.pack(anchor=tk.W)
        
        excel_frame = ttk.Frame(main_frame)
        excel_frame.pack(fill=tk.X, pady=5)
        
        self.excel_entry = ttk.Entry(excel_frame, textvariable=self.excel_path_var, state='readonly')
        self.excel_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 5))
        
        excel_button = ttk.Button(
            excel_frame,
            text="Seleccionar Excel",
            command=self.select_excel_file
        )
        excel_button.pack(side=tk.RIGHT)
        
        # Input PDF file
        self.input_path = StringVar()
        input_label = ttk.Label(main_frame, text="Archivo PDF:", font=('Helvetica', 10))
        input_label.pack(anchor=tk.W)
        
        input_frame = ttk.Frame(main_frame)
        input_frame.pack(fill=tk.X, pady=5)
        
        self.input_entry = ttk.Entry(input_frame, textvariable=self.input_path, state='readonly')
        self.input_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 5))
        
        input_button = ttk.Button(
            input_frame,
            text="Seleccionar PDF",
            command=self.select_input_file
        )
        input_button.pack(side=tk.RIGHT)
        
        # Output directory
        self.output_path = StringVar()
        output_label = ttk.Label(main_frame, text="Carpeta de salida:", font=('Helvetica', 10))
        output_label.pack(anchor=tk.W, pady=(10, 0))
        
        output_frame = ttk.Frame(main_frame)
        output_frame.pack(fill=tk.X, pady=5)
        
        self.output_entry = ttk.Entry(output_frame, textvariable=self.output_path, state='readonly')
        self.output_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 5))
        
        output_button = ttk.Button(
            output_frame,
            text="Seleccionar carpeta",
            command=self.select_output_directory
        )
        output_button.pack(side=tk.RIGHT)
        
        # Process button
        self.process_button = ttk.Button(
            main_frame,
            text="Procesar PDF",
            command=self.start_processing,
        )
        self.process_button.pack(pady=20)

    def select_input_file(self):
        try:
            file_types = [('PDF files', '*.pdf')]
            filename = filedialog.askopenfilename(
                title="Seleccionar archivo PDF",
                filetypes=file_types,
                parent=self.root
            )
            
            if filename:
                self.input_path.set(filename)
        except Exception as e:
            messagebox.showerror("Error", f"Error al seleccionar el archivo: {str(e)}")

    def select_excel_file(self):
        try:
            file_types = [('Excel files', '*.xlsx')]
            filename = filedialog.askopenfilename(
                title="Seleccionar archivo Excel",
                filetypes=file_types,
                parent=self.root
            )
            
            if filename:
                extension = os.path.splitext(filename)[1].lower()
                if extension != '.xlsx':
                    messagebox.showerror("Error", "Solo se permiten archivos con extensión .xlsx")
                    return

                self.excel_path_var.set(filename)
                try:
                    wb = openpyxl.load_workbook(filename, data_only=True)
                    sheet = wb["PROFORMA"]
                    
                    # Buscar la columna CLIENTE en las primeras 20 filas
                    cliente_col = None
                    cliente_row = None
                    for row in range(1, 21):  # Buscar en las primeras 20 filas
                        for col in range(1, sheet.max_column + 1):
                            cell_value = str(sheet.cell(row=row, column=col).value or '').upper()
                            if 'CLIENTE' in cell_value:
                                cliente_col = col
                                cliente_row = row
                                break
                        if cliente_col:
                            break
                    
                    if not cliente_col:
                        raise ValueError("No se encontró la columna CLIENTE en las primeras 20 filas")
                    
                    # Buscar la columna SUBPARTIDA
                    subpartida_col = None
                    for col in range(1, sheet.max_column + 1):
                        cell_value = str(sheet.cell(row=cliente_row, column=col).value or '').upper()
                        if 'SUBPARTIDA' in cell_value:
                            subpartida_col = col
                            break

                    # Buscar la columna descripcion
                    description_col = None
                    for col in range(1, sheet.max_column + 1):
                        cell_value = str(sheet.cell(row=cliente_row, column=col).value or '').upper()
                        if 'DESCRIPCION DECLARADA - PREINSPECCION' in cell_value:
                            description_col = col
                            break

                    # Obtener valores únicos de clientes y sus subpartidas
                    self.clientes_info = {}
                    for row in range(cliente_row + 1, sheet.max_row + 1):
                        cliente = sheet.cell(row=row, column=cliente_col).value
                        subpartida = sheet.cell(row=row, column=subpartida_col).value if subpartida_col else None
                        description = sheet.cell(row=row, column=description_col).value if description_col else None
                        
                        if cliente and isinstance(cliente, str):
                            cliente = cliente.strip()
                            if cliente not in self.clientes_info:
                                self.clientes_info[cliente] = []
                            
                            if subpartida:
                                # Limpiar la subpartida para obtener solo números
                                subpartida_num = re.sub(r'[^0-9]', '', str(subpartida))
                                if subpartida_num:
                                    self.clientes_info[cliente].append({
                                        'numero': subpartida_num,
                                        'descripcion': description  # Puedes agregar la descripción si es necesario
                                    })
                    
                    print("\nClientes y subpartidas detectados:")
                    for cliente, info in self.clientes_info.items():
                        print(f"\nCliente: {cliente}")
                        for subpartida_info in info:
                            print(f"  Subpartida: {subpartida_info['numero']}")
                            print(f"  descripcion: {subpartida_info['descripcion']}")
                    
                except Exception as e:
                    messagebox.showerror("Error", f"Error al leer el archivo Excel: {str(e)}")
                    self.excel_path_var.set("")
                    self.excel_data = None
        except Exception as e:
            messagebox.showerror("Error", f"Error al seleccionar el archivo: {str(e)}")
            
    def map_columns(self, data, header_row, required_columns):
        """
        Mapea las columnas encontradas con las columnas requeridas.
        """
        if header_row is None:
            return None
            
        headers = [str(col).upper().strip() for col in data[header_row]]
        column_mapping = {}
        
        print("\nMapeando columnas:")
        print(f"Headers encontrados: {headers}")
        
        for req_col in required_columns:
            req_col_upper = req_col.upper()
            found = False
            
            for idx, header in enumerate(headers):
                if req_col_upper in header:
                    column_mapping[req_col] = idx
                    print(f"Columna '{req_col}' mapeada al índice {idx}")
                    found = True
                    break
                    
            if not found:
                print(f"No se encontró mapeo para la columna '{req_col}'")
        
        return column_mapping if len(column_mapping) == len(required_columns) else None
    
    def detect_clients(self):
        """
        Detecta los clientes y sus subpartidas desde el Excel.
        """
        if self.excel_data is not None and not self.excel_data.empty:
            # Asegurarse de que las columnas existan
            required_cols = ['CLIENTE', 'SUBPARTIDA', 'DESCRIPCION DECLARADA - PREINSPECCION']
            if not all(col in self.excel_data.columns for col in required_cols):
                print("Columnas requeridas no encontradas:", required_cols)
                return

            # Limpiar y procesar los datos
            self.excel_data['CLIENTE'] = self.excel_data['CLIENTE'].astype(str).apply(
                lambda x: x.strip() if not pd.isna(x) and x != 'nan' else ''
            )
            
            # Filtrar filas con cliente válido
            valid_data = self.excel_data[self.excel_data['CLIENTE'].str.len() > 0]
            
            # Reiniciar el diccionario de clientes
            self.clientes_info = {}
            
            # Procesar cada fila válida
            for _, row in valid_data.iterrows():
                cliente = row['CLIENTE']
                subpartida = re.sub(r'[^0-9]', '', str(row['SUBPARTIDA']))
                descripcion = str(row.get('DESCRIPCION DECLARADA - PREINSPECCION', ''))
                
                if cliente and subpartida:
                    if cliente not in self.clientes_info:
                        self.clientes_info[cliente] = []
                    
                    self.clientes_info[cliente].append({
                        'cliente': cliente,
                        'numero': subpartida,
                        'descripcion': descripcion
                    })
            
            # Imprimir información de depuración
            print("\nClientes detectados:")
            for cliente, info in self.clientes_info.items():
                print(f"\nCliente: {cliente}")
                for subpartida_info in info:
                    print(f"  Subpartida: {subpartida_info['numero']}")
                    print(f"  Descripción: {subpartida_info['descripcion']}")

    def process_excel_data(self):
        try:
            if self.clientes_info:
                for cliente, subpartidas in self.clientes_info.items():
                    print(f"Procesando cliente: {cliente}")
                    for subpartida in subpartidas:
                        print(subpartida)
                    self.create_client_pdf(cliente, subpartidas)
            else:
                print("No hay información de clientes.")
        except Exception as e:
            print(f"Error: {e}")
    
    def select_output_directory(self):
        try:
            directory = filedialog.askdirectory(
                title="Seleccionar carpeta de salida",
                parent=self.root
            )
            
            if directory:
                self.output_path.set(directory)
                # Crear la carpeta declaraciones_separadas
                separated_dir = os.path.join(directory, "declaraciones_separadas")
                if not os.path.exists(separated_dir):
                    os.makedirs(separated_dir)
                # Crear la carpeta separados_por_cliente
                clients_dir = os.path.join(directory, "separados_por_cliente")
                if not os.path.exists(clients_dir):
                    os.makedirs(clients_dir)
        except Exception as e:
            messagebox.showerror("Error", f"Error al seleccionar la carpeta: {str(e)}")
    
    def create_client_pdf(self, cliente, subpartidas):
        """
        Crea un PDF para cada cliente combinando sus subpartidas correspondientes.
        
        Args:
            cliente (str): Nombre del cliente
            subpartidas (list): Lista de diccionarios con información de subpartidas del cliente
        """
        try:
            merger = PdfMerger()
            pdfs_added = False
            separated_dir = os.path.join(self.output_path.get(), "declaraciones_separadas")
            clients_dir = os.path.join(self.output_path.get(), "separados_por_cliente")
            
            print(f"\nProcesando cliente: {cliente}")
            print(f"Buscando archivos en: {separated_dir}")
            
            for subpartida_info in subpartidas:
                subpartida = subpartida_info['numero']
                print(f"Buscando archivos para subpartida: {subpartida}")
                
                # Buscar todos los archivos relacionados con esta subpartida
                subpartida_pattern = f"subpartida_{subpartida}*.pdf"
                matching_files = glob.glob(os.path.join(separated_dir, subpartida_pattern))
                
                if matching_files:
                    print(f"Archivos encontrados para subpartida {subpartida}:")
                    for pdf_file in matching_files:
                        print(f"- {pdf_file}")
                        merger.append(pdf_file)
                        pdfs_added = True
                else:
                    print(f"No se encontraron archivos para subpartida {subpartida}")
            
            if pdfs_added:
                # Crear nombre de archivo válido para Windows
                cliente_filename = re.sub(r'[<>:"/\\|?*]', '_', str(cliente))
                output_path = os.path.join(clients_dir, f"{cliente_filename}.pdf")
                
                # Asegurarse de que el directorio existe
                os.makedirs(clients_dir, exist_ok=True)
                
                # Guardar el PDF combinado
                merger.write(output_path)
                print(f"PDF creado para cliente {cliente}: {output_path}")
            else:
                print(f"No se encontraron PDFs para combinar para el cliente {cliente}")
            
            merger.close()
            
        except Exception as e:
            print(f"Error al crear PDF para cliente {cliente}: {str(e)}")
            messagebox.showerror("Error", f"Error al crear PDF para cliente {cliente}: {str(e)}")

    def show_pdf_selection_window(self, pdf1_path, pdf2_path, descripcion):
        selection_window = tk.Toplevel(self.root)
        selection_window.title("Seleccionar PDF")
        selection_window.geometry("800x600")
        
        # Mostrar descripción
        desc_label = ttk.Label(selection_window, text=f"Descripción: {descripcion}", wraplength=700)
        desc_label.pack(pady=10)
        
        # Frame para los PDFs
        pdf_frame = ttk.Frame(selection_window)
        pdf_frame.pack(fill=tk.BOTH, expand=True)
        
        # Variables para almacenar la selección
        selected_pdf = tk.StringVar()
        
        # Crear dos canvas para mostrar los PDFs
        left_frame = ttk.LabelFrame(pdf_frame, text="PDF Original")
        left_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=5)
        
        right_frame = ttk.LabelFrame(pdf_frame, text="PDF Copia")
        right_frame.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True, padx=5)
        
        # Botones de selección
        ttk.Radiobutton(selection_window, text="Seleccionar Original", 
                       variable=selected_pdf, value=pdf1_path).pack(pady=5)
        ttk.Radiobutton(selection_window, text="Seleccionar Copia", 
                       variable=selected_pdf, value=pdf2_path).pack(pady=5)
        
        # Botón de confirmación
        def confirm_selection():
            selection_window.selected_pdf = selected_pdf.get()
            selection_window.destroy()
            
        ttk.Button(selection_window, text="Confirmar", 
                  command=confirm_selection).pack(pady=10)
        
        # Esperar a que se cierre la ventana
        self.root.wait_window(selection_window)
        
        return getattr(selection_window, 'selected_pdf', None)

    def start_processing(self):
        if not self.validate_inputs():
            return
            
        try:
            self.root.withdraw()
            PreviewWindow(self.input_path.get(), self.output_path.get(), self)
        except Exception as e:
            self.root.deiconify()
            messagebox.showerror("Error", f"Error al iniciar el procesamiento: {str(e)}")

    def validate_inputs(self):
        if not all([self.input_path.get(), self.output_path.get(), self.excel_path_var.get()]):
            messagebox.showerror("Error", "Por favor seleccione todos los archivos necesarios")
            return False
        return True

class PDFPreviewDialog(tk.Toplevel):
    def __init__(self, parent, pdfs, cliente, descriptions, numero_subpartida, descripcion_subpartida, title="Seleccionar PDF"):
        super().__init__(parent)
        self.title(title)
        self.numero_subpartida = numero_subpartida  # Número de subpartida
        self.descripcion_subpartida = descripcion_subpartida  # Descripción de la subpartida
        self.nombre_cliente = cliente
        self.selected_pdfs = []
        self.selected_variable = tk.BooleanVar(value=False)  # Variable de control para coordinar selección
        self.setup_window()
        self.create_widgets(pdfs, descriptions)
        
        # Evitar que se cierre la ventana accidentalmente sin selección
        self.protocol("WM_DELETE_WINDOW", self.on_close)

    def setup_window(self):
        # Set window size to 80% of screen size
        # screen_width = self.winfo_screenwidth()
        # screen_height = self.winfo_screenheight()
        # # window_width = int(screen_width * 0.8)
        # # window_height = int(screen_height * 0.8)

        # x = (screen_width - window_width) // 2
        # y = (screen_height - window_height) // 2
        # self.geometry(f"{window_width}x{window_height}+{x}+{y}")
        self.attributes('-fullscreen', True)

        
    def create_widgets(self, pdfs, descriptions):
        # Etiqueta con número y descripción de la subpartida
        # Frame para contener todas las etiquetas
        text_frame = ttk.Frame(self)
        text_frame.pack(pady=20)
        
        # Título con texto más grande
        title_label = ttk.Label(
            text_frame,
            text=f"Para el cliente {self.nombre_cliente}",
            font=('Helvetica', 16, 'bold')  # Fuente más grande para el título
        )
        title_label.pack(pady=(0, 10))
        
        # Mensaje sobre archivos encontrados
        files_label = ttk.Label(
            text_frame,
            text="Se encontraron 2 archivos o más para esta subpartida.",
            font=('Helvetica', 12)
        )
        files_label.pack(pady=2)
        
        # Número de subpartida con fondo amarillo
        subpartida_frame = ttk.Frame(text_frame)
        subpartida_frame.pack(pady=2, fill='x')
        subpartida_label = tk.Label(  # Usar tk.Label en lugar de ttk.Label para poder cambiar el color
            subpartida_frame,
            text=f"Subpartida número: {self.numero_subpartida}",
            font=('Helvetica', 12),
            bg='yellow',  # Fondo amarillo
            fg='black'    # Texto negro
        )
        subpartida_label.pack()
        
        # Descripción con fondo blanco
        descripcion_label = tk.Label(  # Usar tk.Label en lugar de ttk.Label para poder cambiar el color
            text_frame,
            text=f"Descripción de esta subpartida: {self.descripcion_subpartida}",
            font=('Helvetica', 12),
            bg='white',
            fg='black',
            wraplength=800  # Para que el texto se ajuste si es muy largo
        )
        descripcion_label.pack(pady=2)
        
        # Pregunta final
        question_label = ttk.Label(
            text_frame,
            text="¿Desea guardar todos o seleccionar uno específico?",
            font=('Helvetica', 12)
        )
        question_label.pack(pady=(10, 0))
        
        # Descriptions
        desc_frame = ttk.LabelFrame(self, text="Descripciones encontradas")
        desc_frame.pack(fill=tk.X, padx=20, pady=10)
        for desc in descriptions:
            ttk.Label(desc_frame, text=desc, wraplength=800).pack(pady=5)
            
        # PDF Previews
        preview_frame = ttk.Frame(self)
        preview_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=10)
        
        # Create preview containers for each PDF
        self.preview_containers = []
        for i, pdf_path in enumerate(pdfs):
            container = PDFPreviewContainer(preview_frame, pdf_path)
            container.grid(row=0, column=i, padx=10, pady=10, sticky="nsew")
            self.preview_containers.append(container)
            preview_frame.grid_columnconfigure(i, weight=1)
            
        # Buttons
        btn_frame = ttk.Frame(self)
        btn_frame.pack(pady=20)
        
        ttk.Button(
            btn_frame,
            text="Guardar todos", 
            command=lambda: self.finish_selection(pdfs)
        ).pack(side=tk.LEFT, padx=10)
        
        ttk.Button(
            btn_frame,
            text="Guardar seleccionados", 
            command=self.save_selected
        ).pack(side=tk.LEFT, padx=10)
                  
    def save_selected(self):
        self.selected_pdfs = [
            container.pdf_path for container in self.preview_containers 
            if container.is_selected
        ]
        self.selected_variable.set(True)  # Indicar que se completó la selección
        self.destroy()
        
    def finish_selection(self, all_pdfs=None):
        self.selected_pdfs = all_pdfs if all_pdfs else self.selected_pdfs
        self.selected_variable.set(True)  # Indicar que se completó la selección
        self.destroy()

    def on_close(self):
        """Método para manejar el cierre de la ventana sin selección explícita."""
        self.selected_variable.set(True)  # Indicar que la ventana se cerró
        self.destroy()


class PDFPreviewContainer(ttk.Frame):
    def __init__(self, parent, pdf_path):
        super().__init__(parent)
        self.pdf_path = pdf_path
        self.is_selected = False
        self.current_page = 0
        self.total_pages = 0
        self.zoom_factor = 2.0  # Inicializar el factor de zoom
        self.setup_preview()

    def setup_preview(self):
        self.config(relief="solid", borderwidth=1)
        
        # Frame para el canvas y barras de desplazamiento
        canvas_frame = ttk.Frame(self)
        canvas_frame.pack(fill=tk.BOTH, expand=True, pady=10)
        
        # Canvas para la previsualización del PDF con altura fija de 900px
        self.canvas = tk.Canvas(canvas_frame, bg="white", bd=0, highlightthickness=0, height=900)
        self.canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        # Scrollbars
        self.v_scrollbar = ttk.Scrollbar(canvas_frame, orient="vertical", command=self.canvas.yview)
        self.v_scrollbar.pack(side=tk.RIGHT, fill="y")
        
        self.h_scrollbar = ttk.Scrollbar(self, orient="horizontal", command=self.canvas.xview)
        self.h_scrollbar.pack(side=tk.BOTTOM, fill="x")
        
        # Configurar los scrollbars
        self.canvas.config(yscrollcommand=self.v_scrollbar.set, xscrollcommand=self.h_scrollbar.set)
        
        # Botones de navegación
        nav_frame = ttk.Frame(self)
        nav_frame.pack(fill=tk.X, pady=5)
        
        ttk.Button(nav_frame, text="←", command=self.prev_page).pack(side=tk.LEFT, padx=5)
        self.page_label = ttk.Label(nav_frame, text="Pagina: 1/1")
        self.page_label.pack(side=tk.LEFT, expand=True)
        ttk.Button(nav_frame, text="→", command=self.next_page).pack(side=tk.RIGHT, padx=5)
        
        # Botones de zoom
        zoom_frame = ttk.Frame(self)
        zoom_frame.pack(fill=tk.X, pady=5)
        
        ttk.Button(zoom_frame, text="Acercar 🔎+", command=self.zoom_in).pack(side=tk.LEFT, padx=5)
        ttk.Button(zoom_frame, text="Alejar 🔍-", command=self.zoom_out).pack(side=tk.RIGHT, padx=5)
        
        # Cargar el PDF
        self.load_pdf()
        
        # Bind click event
        self.canvas.bind("<Button-1>", self.toggle_selection)

    def load_pdf(self):
        try:
            self.pdf_document = fitz.open(self.pdf_path)
            self.total_pages = len(self.pdf_document)
            self.update_preview()
        except Exception as e:
            print(f"Error loading PDF: {e}")

    def update_preview(self):
        if not hasattr(self, 'pdf_document'):
            return
        
        page = self.pdf_document[self.current_page]
        
        # Usar el factor de zoom para ajustar la escala de la imagen
        zoom_matrix = fitz.Matrix(self.zoom_factor, self.zoom_factor)
        pix = page.get_pixmap(matrix=zoom_matrix)
        img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
        self.preview_image = ImageTk.PhotoImage(image=img)
        
        self.canvas.delete("all")  # Eliminar elementos previos en el lienzo
        self.canvas.create_image(0, 0, anchor=tk.NW, image=self.preview_image)

        # Crear el borde verde solo cuando se selecciona, y hacerlo un poco más grueso
        if self.is_selected:
            self.canvas.create_rectangle(
                0, 0, pix.width, pix.height, outline="green", width=5
            )
        
        self.page_label.config(text=f"Pagina: {self.current_page + 1}/{self.total_pages}")
        
        # Configurar las barras de desplazamiento
        self.canvas.config(scrollregion=self.canvas.bbox("all"))

        # **Nuevo código**: Centrar el scroll en la mitad
        self.center_scroll()

    def center_scroll(self):
        """Ajustar el scroll en la mitad del lienzo."""
        bbox = self.canvas.bbox("all")  # Obtener el área de la previsualización
        if bbox:
            # Calcular las posiciones iniciales de scroll en x y y para centrar
            canvas_width = bbox[2] - bbox[0]
            canvas_height = bbox[3] - bbox[1]
            
            center_x = (canvas_width - self.canvas.winfo_width()) // 2
            center_y = (canvas_height - self.canvas.winfo_height()) // 2
            
            # Mover las barras de desplazamiento a la mitad
            self.canvas.yview_moveto(center_y / canvas_height)
            self.canvas.xview_moveto(center_x / canvas_width)
        
    def toggle_selection(self, event=None):
        """Alterna entre seleccionar y deseleccionar."""
        if self.is_selected:
            self.is_selected = False
        else:
            self.is_selected = True
        
        # Después de alternar la selección, actualizamos la vista
        self.update_preview()

    def next_page(self):
        if self.current_page < self.total_pages - 1:
            self.current_page += 1
            self.update_preview()

    def prev_page(self):
        if self.current_page > 0:
            self.current_page -= 1
            self.update_preview()

    def zoom_in(self):
        """Incrementar el factor de zoom."""
        self.zoom_factor *= 1.2
        self.update_preview()

    def zoom_out(self):
        """Disminuir el factor de zoom."""
        self.zoom_factor /= 1.2
        self.update_preview()

# Now let's modify the PreviewWindow class to handle automatic subpartida detection:
class PreviewWindow:
    def __init__(self, input_path, output_path, parent_window):
        self.root = tk.Toplevel()
        self.input_path = input_path
        self.output_path = output_path
        self.parent_window = parent_window
        self.current_page = 0
        self.pdf_document = None
        self.preview_image = None
        self.page_data = []
        self.zoom_factor = 3.5
        self.last_selection_coords = None
        self.zoom_locked = False
        self.ocr_lang = 'eng'
        # Añadir diccionario para agrupar descripciones
        self.descriptions_by_subpartida = {}
        # Añadir tolerancia para coordenadas
        self.coord_tolerance = 20  # píxeles de tolerancia
        if platform.system() == 'Windows':
            pytesseract.pytesseract.tesseract_cmd = r'C:\Program Files\Tesseract-OCR\tesseract.exe'
        
        self.show_initial_zoom_message()
        self.setup_window()
        self.create_widgets()
        self.load_pdf()
        self.process_excel_descriptions()

    def process_excel_descriptions(self):
        """
        Agrupa las descripciones del Excel por subpartida
        """
        for cliente, subpartidas in self.parent_window.clientes_info.items():
            for subpartida_info in subpartidas:
                subpartida = re.sub(r'[^0-9]', '', subpartida_info['numero'])
                descripcion = subpartida_info.get('descripcion', '')
                
                if subpartida not in self.descriptions_by_subpartida:
                    self.descriptions_by_subpartida[subpartida] = set()
                
                if descripcion:
                    self.descriptions_by_subpartida[subpartida].add(descripcion)
        
        # Imprimir en consola las descripciones agrupadas
        print("\nDescripciones agrupadas por subpartida:")
        for subpartida, descriptions in self.descriptions_by_subpartida.items():
            print(f"\nSubpartida {subpartida}:")
            for desc in descriptions:
                print(f"  - {desc}")
    def show_initial_zoom_message(self):
            messagebox.showinfo(
                "Importante",
                "Por favor, ajuste el zoom para ver claramente el número de subpartida "
                "antes de hacer la primera selección. Una vez seleccionada la primera "
                "subpartida, el zoom quedará bloqueado para mantener las coordenadas "
                "consistentes."
            )
    
    def lock_zoom(self):
        self.zoom_locked = True
        self.zoom_in_button.state(['disabled'])
        self.zoom_out_button.state(['disabled'])
    
    def handle_selection(self, coords):
        if self.current_page == 0 or (self.page_data and self.page_data[-1]['type'] == 'e'):
            # Primera página o cambio de espaldar a principal
            self.last_selection_coords = coords
            if not self.zoom_locked:
                self.lock_zoom()
        elif self.last_selection_coords:
            # Usar las coordenadas guardadas para detección automática
            self.auto_detect_subpartida(self.last_selection_coords)

    # TODO: auto detect subpartida
    def auto_detect_subpartida(self, coords=None):
        # Si no se pasan coordenadas, usa las predeterminadas
        if coords is None and self.last_selection_coords is None:
            print("No se han proporcionado coordenadas para la detección.")
            return
        coords = coords or self.last_selection_coords

        print(f"\nDetección automática de subpartida: coordenadas {coords}")
        # Deshabilitar el botón
        self.zoom_in_button.config(state="disabled")
        self.zoom_out_button.config(state="disabled")
        try:
            x0, y0, x1, y1 = coords
            page = self.pdf_document[self.current_page]
            pix = page.get_pixmap(matrix=fitz.Matrix(self.zoom_factor, self.zoom_factor))
            img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
            
            # Ampliar el área de selección con la tolerancia
            selection = img.crop((
                max(0, min(x0, x1) - self.coord_tolerance),
                max(0, min(y0, y1) - self.coord_tolerance),
                min(pix.width, max(x0, x1) + self.coord_tolerance),
                min(pix.height, max(y0, y1) + self.coord_tolerance)
            ))
            
            processed_image = self.preprocess_image(selection)
            text = pytesseract.image_to_string(processed_image, config='--psm 6')
            
            # Buscar números con formato de subpartida
            matches = re.findall(r'\d{4,}(?:\.\d+)*', text)
            if matches:
                longest_match = max(matches, key=len)
                self.subpartida_var.set(longest_match)
                self.tipo_var.set('p')
                print(f"Subpartida detectada: {longest_match}")
                
                subpartida_base = re.sub(r'[^0-9]', '', longest_match)
                if subpartida_base in self.descriptions_by_subpartida:
                    print(f"\nDescripciones encontradas para subpartida {longest_match}:")
                    for desc in self.descriptions_by_subpartida[subpartida_base]:
                        print(f"  - {desc}")
            else:
                print("No se detectó ningún número de subpartida en la selección")
                self.tipo_var.set('e')
            
        except Exception as e:
            print(f"Error en detección automática: {str(e)}")
            self.tipo_var.set('e')
    def on_mouse_wheel(self, event):
        if not self.zoom_locked:
            if event.delta > 0:
                self.canvas.yview_scroll(-1, "units")
            else:
                self.canvas.yview_scroll(1, "units")
    def preprocess_image(self, image):
        # Convert PIL Image to OpenCV format
        opencv_image = cv2.cvtColor(np.array(image), cv2.COLOR_RGB2BGR)
        
        # Convert to grayscale
        gray = cv2.cvtColor(opencv_image, cv2.COLOR_BGR2GRAY)
        
        # Apply adaptive thresholding
        thresh = cv2.adaptiveThreshold(
            gray, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C, cv2.THRESH_BINARY, 11, 2
        )
        
        return thresh

    def extract_text_from_roi(self, image, roi_coords, config):
        # Extract region of interest (ROI)
        x1, y1, x2, y2 = roi_coords
        roi = image[y1:y2, x1:x2]
        
        # Extract text with Tesseract
        text = pytesseract.image_to_string(roi, config=config)
        
        return text

    def check_tesseract_installation(self):
        try:
            pytesseract.get_tesseract_version()
            return True
        except Exception:
            return False

    def extract_numbers_from_text(self, text):
        match = re.search(r'\d+', text)
        if match:
            return match.group()
        return None

    def setup_window(self):
        self.root.title("Previsualización y Edición")
        # Iniciar en pantalla completa
        self.root.attributes('-fullscreen', True)
        # Añadir accesos rápidos
        self.root.bind('<Escape>', lambda e: self.toggle_fullscreen())
        self.root.bind('<Control-plus>', lambda e: self.zoom_in())
        self.root.bind('<Control-minus>', lambda e: self.zoom_out())
        
    def toggle_fullscreen(self):
        if self.root.attributes('-fullscreen'):
            self.root.attributes('-fullscreen', False)
        else:
            self.root.attributes('-fullscreen', True)
            
    def create_widgets(self):
        main_frame = ttk.Frame(self.root, padding="20")
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # Frame superior para controles
        control_frame = ttk.Frame(main_frame)
        control_frame.pack(fill=tk.X, pady=(0, 10))
        
        # Navegación
        nav_frame = ttk.Frame(control_frame)
        nav_frame.pack(side=tk.LEFT)
        
        self.prev_button = ttk.Button(nav_frame, text="← Anterior", command=self.prev_page)
        self.prev_button.pack(side=tk.LEFT, padx=5)
        
        self.page_label = ttk.Label(nav_frame, text="Página: 0/0")
        self.page_label.pack(side=tk.LEFT, padx=20)
        
        self.next_button = ttk.Button(nav_frame, text="Siguiente →", command=self.save_and_next)
        self.next_button.pack(side=tk.LEFT, padx=5)
        
        # Frame para número de subpartida
        subpartida_frame = ttk.Frame(control_frame)
        subpartida_frame.pack(side=tk.LEFT, padx=20)
        
        self.subpartida_var = StringVar()
        ttk.Label(subpartida_frame, text="Número de Subpartida:", font=('Helvetica', 10)).pack(side=tk.LEFT, padx=5)
        self.subpartida_entry = ttk.Entry(subpartida_frame, textvariable=self.subpartida_var, width=20)
        self.subpartida_entry.pack(side=tk.LEFT, padx=5)
        
        # Frame para tipo de página (botones radio)
        tipo_frame = ttk.Frame(control_frame)
        tipo_frame.pack(side=tk.LEFT, padx=20)
        
        self.tipo_var = StringVar(value='p')  # Default a principal
        ttk.Label(tipo_frame, text="Tipo de Página:", font=('Helvetica', 10)).pack(side=tk.LEFT, padx=5)
        
        self.radio_principal = ttk.Radiobutton(tipo_frame, text="Principal", value='p', variable=self.tipo_var)
        self.radio_principal.pack(side=tk.LEFT, padx=2)
        
        self.radio_espaldar = ttk.Radiobutton(tipo_frame, text="Espaldar", value='e', variable=self.tipo_var)
        self.radio_espaldar.pack(side=tk.LEFT, padx=2)
        
         # Mejorar los botones de zoom
        zoom_frame = ttk.Frame(control_frame)
        zoom_frame.pack(side=tk.RIGHT, padx=20)
        
        # Opción 1: Usando caracteres Unicode
        ttk.Label(zoom_frame, text="Zoom:").pack(side=tk.LEFT, padx=5)
        self.zoom_in_button = ttk.Button(zoom_frame, text="🔍+ (Ctrl +)", command=self.zoom_in)
        self.zoom_in_button.pack(side=tk.LEFT, padx=2)

        self.zoom_out_button = ttk.Button(zoom_frame, text="🔍- (Ctrl -)", command=self.zoom_out)
        self.zoom_out_button.pack(side=tk.LEFT, padx=2)
        
        # Añadir etiqueta para mostrar el factor de zoom actual
        self.zoom_label = ttk.Label(zoom_frame, text="150%")
        self.zoom_label.pack(side=tk.LEFT, padx=5)
        
        # Botones de acción
        action_frame = ttk.Frame(control_frame)
        action_frame.pack(side=tk.RIGHT)
        
        self.save_button = ttk.Button(action_frame, text="Guardar Todo", command=self.save_pdfs)
        self.save_button.pack(side=tk.RIGHT, padx=5)
        
        # Configurar teclas rápidas
        self.root.bind('p', lambda e: self.tipo_var.set('p'))
        self.root.bind('e', lambda e: self.tipo_var.set('e'))
        
        # Canvas y scrollbars (mismo código que antes)
        preview_frame = ttk.Frame(main_frame)
        preview_frame.pack(fill=tk.BOTH, expand=True)
        
        self.v_scrollbar = ttk.Scrollbar(preview_frame, orient="vertical")
        self.h_scrollbar = ttk.Scrollbar(preview_frame, orient="horizontal")
        
        self.canvas = Canvas(
            preview_frame,
            bg='white',
            yscrollcommand=self.v_scrollbar.set,
            xscrollcommand=self.h_scrollbar.set,
            cursor="pencil"  # Cambiar el cursor a lápiz
        )
        
        self.v_scrollbar.config(command=self.canvas.yview)
        self.h_scrollbar.config(command=self.canvas.xview)
        
        self.v_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.h_scrollbar.pack(side=tk.BOTTOM, fill=tk.X)
        self.canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        
        # Configurar eventos del mouse para selección de texto
        self.canvas.bind("<ButtonPress-1>", self.on_button_press)
        self.canvas.bind("<B1-Motion>", self.on_mouse_drag)
        self.canvas.bind("<ButtonRelease-1>", self.on_button_release)
    def zoom_in(self):
        if not self.zoom_locked:
            self.zoom_factor += 0.3
            self.update_zoom_label()
            self.update_page_display()

    def zoom_out(self):
        if not self.zoom_locked:
            self.zoom_factor = max(0.3, self.zoom_factor - 0.3)
            self.update_zoom_label()
            self.update_page_display()
    
    def update_zoom_label(self):
        zoom_percentage = int(self.zoom_factor * 100)
        self.zoom_label.config(text=f"{zoom_percentage}%")

    def save_and_next(self):
        # Guardar datos de la página actual
        page_info = {
            'page_number': self.current_page,
            'type': self.tipo_var.get(),
            'subpartida': self.subpartida_var.get() if self.tipo_var.get() == 'p' else None
        }
        
        # Si es una página nueva, añadirla; si existe, actualizarla
        if self.current_page >= len(self.page_data):
            self.page_data.append(page_info)
        else:
            self.page_data[self.current_page] = page_info
        
        # Ir a la siguiente página
        if self.current_page < self.pdf_document.page_count - 1:
            self.current_page += 1
            self.update_page_display()
            
            # Cargar datos guardados si existen
            if self.current_page < len(self.page_data):
                saved_data = self.page_data[self.current_page]
                self.tipo_var.set(saved_data['type'])
                if saved_data['subpartida']:
                    self.subpartida_var.set(saved_data['subpartida'])
                else:
                    self.subpartida_var.set('')

    def load_pdf(self):
        try:
            self.pdf_document = fitz.open(self.input_path)
            self.update_page_display()
        except Exception as e:
            messagebox.showerror("Error", f"Error al cargar el PDF: {str(e)}")

    def get_page_text_ocr(self, page):
        try:
            pix = page.get_pixmap(matrix=fitz.Matrix(2.0, 2.0))
            img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
            
            # Use default language if Spanish not available
            try:
                text = pytesseract.image_to_string(img, lang='eng')
            except pytesseract.TesseractError:
                text = pytesseract.image_to_string(img)
            
            return text
        except Exception as e:
            print(f"OCR Error: {str(e)}")
            return ""

    def detect_subpartidas_and_backups(self):
        if not self.pdf_document:
            return

        patterns = [
            r'[Ss]ubpartida\s*[Aa]rancelaria\s*[:#]?\s*(\d{4,}(?:\.\d+)?)',
            r'[Ss]ubpartida\s*[:#]?\s*(\d{4,}(?:\.\d+)?)',
            r'SUBPARTIDA\s+ARANCELARIA\s*[:#]?\s*(\d{4,}(?:\.\d+)?)',
            r'(?:^|\s)(\d{4}\.\d{2}\.\d{2})(?:\s|$)',
            r'(?<=subpartida\s)(\d{4,}(?:\.\d+)?)',
        ]

        for page_num in range(self.pdf_document.page_count):
            page = self.pdf_document[page_num]
            # Use OCR instead of direct text extraction
            text = self.get_page_text_ocr(page)
            print(f"\nTexto extraído por OCR de la página {page_num + 1}:")
            print(text)

            found_subpartida = False
            for pattern in patterns:
                matches = re.finditer(pattern, text, re.IGNORECASE)
                for match in matches:
                    potential_subpartida = match.group(1).strip()
                    if self.validate_subpartida_format(potential_subpartida):
                        print(f"Subpartida encontrada: {potential_subpartida}")
                        self.page_data.append({
                            'page_number': page_num,
                            'type': 'p',
                            'subpartida': potential_subpartida
                        })
                        found_subpartida = True
                        break
                if found_subpartida:
                    break

            if not found_subpartida:
                self.page_data.append({
                    'page_number': page_num,
                    'type': 'e',
                    'subpartida': None
                })
                 
    def validate_subpartida_format(self, subpartida):
        """
        Valida el formato de una subpartida
        """
        # Eliminar espacios y puntos extras
        subpartida = subpartida.strip().replace(' ', '')
        
        # Verificar formato básico (debe tener números y opcionalmente puntos)
        if not re.match(r'^\d+(\.\d+)*$', subpartida):
            return False

        # Verificar longitud mínima y máxima
        if len(subpartida.replace('.', '')) < 4 or len(subpartida.replace('.', '')) > 10:
            return False

        # Verificar que cada sección entre puntos sea válida
        parts = subpartida.split('.')
        for part in parts:
            if not (1 <= len(part) <= 4):  # Cada parte debe tener entre 1 y 4 dígitos
                return False

        return True
    
    # TODO: 
    def update_page_display(self):
        if not self.pdf_document:
            return

        try:
            page = self.pdf_document[self.current_page]
            pix = page.get_pixmap(matrix=fitz.Matrix(self.zoom_factor, self.zoom_factor))
            img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
            self.preview_image = ImageTk.PhotoImage(image=img)

            # Actualizar canvas
            self.canvas.delete("all")
            self.canvas.config(scrollregion=(0, 0, pix.width, pix.height))
            self.canvas.create_image(0, 0, anchor=tk.NW, image=self.preview_image)
            
            # Actualizar etiqueta de página
            self.page_label.config(text=f"Página: {self.current_page + 1}/{self.pdf_document.page_count}")

            # Si hay coordenadas de selección previas, usar esas para detectar la subpartida
            if self.last_selection_coords:
                print(f"Usando coordenadas previas para detección: {self.last_selection_coords}")
                self.auto_detect_subpartida(coords=self.last_selection_coords)
            else:
                text = page.get_text("text")
                if isinstance(text, bytes):
                    text = text.decode('utf-8')
                print(f"Texto completo de la página {self.current_page + 1}:")
                print(text)  # Debug

                # Intentar detectar subpartida
                patterns = [
                    r'Subpartida\s+[Aa]rancelaria\s*[:#]?\s*(\d+)',
                    r'SUBPARTIDA\s+ARANCELARIA\s*[:#]?\s*(\d+)',
                    r'Subpartida\s*[:#]?\s*(\d+)',
                ]
                
                found_subpartida = False
                for pattern in patterns:
                    match = re.search(pattern, text)
                    if match:
                        detected_number = match.group(1).strip()
                        print(f"Subpartida detectada: {detected_number}")  # Debug
                        self.subpartida_var.set(detected_number)
                        found_subpartida = True
                        break
                
                if not found_subpartida:
                    print("No se detectó subpartida en esta página")  # Debug

            # Actualizar tipo de página y subpartida
            if self.current_page < len(self.page_data):
                current_data = self.page_data[self.current_page]
                self.tipo_var.set(current_data['type'])
                self.subpartida_var.set(current_data['subpartida'] or '')

            # Asegurar que los widgets estén visibles
            self.subpartida_entry.pack(side=tk.LEFT, padx=5)
            self.radio_principal.pack(side=tk.LEFT, padx=2)
            self.radio_espaldar.pack(side=tk.LEFT, padx=2)
            
            # Configurar el botón de guardar
            if self.current_page == self.pdf_document.page_count - 1:
                self.save_button.pack(side=tk.RIGHT, padx=5)
            else:
                self.save_button.pack_forget()

        except Exception as e:
            error_msg = f"Error en update_page_display: {str(e)}"
            print(error_msg)  # Debug
            messagebox.showerror("Error", f"Error al actualizar la visualización: {str(e)}")

    def next_page(self):
        if self.pdf_document and self.current_page < self.pdf_document.page_count - 1:
            self.current_page += 1
            self.update_page_display()
            
    def prev_page(self):
        if self.pdf_document and self.current_page > 0:
            self.current_page -= 1
            self.update_page_display()
            
    def on_button_press(self, event):
        self.start_x = self.canvas.canvasx(event.x)
        self.start_y = self.canvas.canvasy(event.y)
        self.rect = self.canvas.create_rectangle(
            self.start_x, self.start_y, 
            self.start_x, self.start_y, 
            outline='red', 
            width=2
        )
        
    def on_mouse_drag(self, event):
        cur_x = self.canvas.canvasx(event.x)
        cur_y = self.canvas.canvasy(event.y)
        self.canvas.coords(self.rect, self.start_x, self.start_y, cur_x, cur_y)
        
    # TODO: donde se detecta el numero
    def on_button_release(self, event):
        try:
            coords = self.canvas.coords(self.rect)
            if not coords:
                return
            
            x0, y0, x1, y1 = coords
            page = self.pdf_document[self.current_page]
            
            pix = page.get_pixmap(matrix=fitz.Matrix(self.zoom_factor, self.zoom_factor))
            img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
            
            selection_box = (
                min(x0, x1),
                min(y0, y1),
                max(x0, x1),
                max(y0, y1)
            )
            
            selection = img.crop(selection_box)
            processed_image = self.preprocess_image(selection)
            
            tesseract_config = r'--oem 3 --psm 6'
            height, width = processed_image.shape[:2]
            text = self.extract_text_from_roi(processed_image, (0, 0, width, height), tesseract_config)
            detected_number = self.extract_numbers_from_text(text)
            
            # Print coordinates and detected number
            print(f"Coordinates: {coords}")
            print(f"Detected number: {detected_number}")
            
            if detected_number:
                self.subpartida_var.set(detected_number)
                print(f"Número detectado: {detected_number}")
            
            # Guardar las coordenadas seleccionadas para futuras páginas
            self.last_selection_coords = [x0, y0, x1, y1]
            print(f"Coordenadas almacenadas para futuras detecciones: {self.last_selection_coords}")
            
            self.canvas.delete(self.rect)
            
        except Exception as e:
            print(f"Error en OCR: {str(e)}")
            messagebox.showerror("Error", f"Error al procesar el texto: {str(e)}")

    def show_preview(self, cliente, subpartida, matching_files):
        """
        Muestra una vista previa de los archivos duplicados para que el usuario pueda seleccionar cuál desea incluir.
        """
        preview_window = tk.Toplevel(self.root)
        preview_window.title(f"Seleccionar archivo para {cliente} - Subpartida {subpartida}")
        
        label = ttk.Label(preview_window, text=f"Seleccionar archivo para {cliente} - Subpartida {subpartida}")
        label.pack(pady=10)
        
        file_var = tk.StringVar(value=matching_files[0])
        
        for file in matching_files:
            radio_button = ttk.Radiobutton(preview_window, text=file, variable=file_var, value=file)
            radio_button.pack(anchor=tk.W)
        
        def on_select():
            selected_file = file_var.get()
            self.selected_files[cliente][subpartida] = selected_file
            preview_window.destroy()
        
        select_button = ttk.Button(preview_window, text="Seleccionar", command=on_select)
        select_button.pack(pady=10)
        
        preview_window.transient(self.root)
        preview_window.grab_set()
        self.root.wait_window(preview_window)

    def save_pdfs(self):
        """
        Guarda los PDFs separados por subpartida y luego crea los PDFs por cliente.
        """
        try:
            # Crear directorios necesarios
            separated_dir = os.path.join(self.parent_window.output_path.get(), "declaraciones_separadas")
            clients_dir = os.path.join(self.parent_window.output_path.get(), "separados_por_cliente")
            os.makedirs(separated_dir, exist_ok=True)
            os.makedirs(clients_dir, exist_ok=True)

            # 1. Primero, guardar todas las subpartidas separadas
            input_pdf = PyPDF2.PdfReader(self.input_path)
            current_subpartida = None
            current_pages = []
            subpartida_counts = {}

            # Procesar todas las páginas para crear los PDFs separados
            for page_info in self.page_data:
                if page_info['type'] == 'p':
                    if current_subpartida and current_pages:
                        self.save_single_pdf(input_pdf, current_subpartida, current_pages, subpartida_counts)
                    current_subpartida = page_info['subpartida']
                    current_pages = [page_info['page_number']]
                else:  # Espaldar
                    if current_pages:
                        current_pages.append(page_info['page_number'])

            # Guardar el último grupo
            if current_subpartida and current_pages:
                self.save_single_pdf(input_pdf, current_subpartida, current_pages, subpartida_counts)

            # Inicializar diccionario para archivos seleccionados
            self.selected_files = {cliente: {} for cliente in self.parent_window.clientes_info.keys()}

            # 2. Después de que todos los PDFs separados están guardados, crear los PDFs por cliente
            self.root.withdraw()  # Ocultar la ventana principal
            for cliente, subpartidas in self.parent_window.clientes_info.items():
                merger = PdfMerger()
                pdfs_added = False
                
                print(f"\nProcesando cliente: {cliente}")
                subpartida_seen = set()
                for subpartida_info in subpartidas:
                    subpartida = subpartida_info['numero']
                    descripcion = subpartida_info.get('descripcion', '')
                    print(f"descripcion en  save PDFS method: {descripcion}")
                    # Normalizar el número de subpartida para la búsqueda
                    subpartida_base = re.sub(r'[^0-9]', '', subpartida)
                    if subpartida_base.endswith('0'):
                        subpartida_base = subpartida_base[:-1]
                    print(f"ARCHIVO BASE ORIGINIAL PARA BUSCAR: {subpartida_base}")
                    
                    if subpartida_base in subpartida_seen:
                        continue
                    subpartida_seen.add(subpartida_base)
                    
                    print(f"Buscando archivos para subpartida {subpartida}")
                    
                    # Buscar la subpartida exacta y sus copias
                    subpartida_pattern = f"subpartida_{subpartida_base}*.pdf"
                    search_path = os.path.join(separated_dir, subpartida_pattern)
                    matching_files = glob.glob(search_path)
                    
                    print(f"Buscando con patrón: {subpartida_pattern}")
                    print(f"URL de búsqueda: {search_path}")
                    print(f"Archivos encontrados: {matching_files}")
                    
                    if matching_files:
                        if len(matching_files) > 1:
                            # Obtener descripciones del Excel
                            descriptions = [
                                subp.get('descripcion', '')
                                for subp in self.parent_window.clientes_info[cliente]
                                if re.sub(r'[^0-9]', '', subp['numero']) == subpartida_base
                            ]
                            
                            dialog = PDFPreviewDialog(
                                self.root,
                                matching_files,
                                cliente,
                                descriptions,
                                numero_subpartida=subpartida,  # Pasar número de subpartida
                                descripcion_subpartida=descripcion,  # Pasar descripción
                                title="Seleccionar PDF para Subpartida"
                            )
                            self.root.wait_window(dialog)
                            
                            selected_files = dialog.selected_pdfs
                            if selected_files:
                                for file in selected_files:
                                    merger.append(file)
                                    pdfs_added = True
                        else:
                            merger.append(matching_files[0])
                            pdfs_added = True

                if pdfs_added:
                    cliente_filename = re.sub(r'[<>:"/\\|?*]', '_', str(cliente))
                    output_path = os.path.join(clients_dir, f"{cliente_filename}.pdf")
                    merger.write(output_path)
                    print(f"PDF creado para cliente {cliente}: {output_path}")
                else:
                    print(f"No se encontraron PDFs para combinar para el cliente {cliente}")
                
                merger.close()

            messagebox.showinfo("Éxito", "PDFs generados correctamente")
    
        except Exception as e:
            # Mostrar error si ocurre algo
            print(f"Error al guardar los PDFs: {str(e)}")
            messagebox.showerror("Error", f"Error al guardar los PDFs: {str(e)}")
        
        finally:
            # Asegurarse de destruir la ventana de previsualización y restaurar la principal
            try:
                self.root.destroy()
            except Exception as e:
                print(f"Error al cerrar la ventana de previsualización: {str(e)}")
            self.parent_window.root.deiconify()
        
    def process_excel_data(self):
        if self.excel_data is None:
            messagebox.showerror("Error", "No se ha cargado un archivo Excel válido")
            return

        try:
            # Procesar por cliente
            for cliente, subpartidas in self.clientes_info.items():
                print(f"Procesando cliente: {cliente}")  # Debug
                print(f"Subpartidas para {cliente}: {subpartidas}")  # Debug
                
                if subpartidas:
                    print(f"Creando PDF para cliente {cliente}")  # Debug
                    self.create_client_pdf(cliente, subpartidas)
                    
            messagebox.showinfo("Éxito", "Proceso completado correctamente")
        except Exception as e:
            messagebox.showerror("Error", f"Error al procesar el archivo Excel: {str(e)}")
    
    def save_single_pdf(self, input_pdf, subpartida, pages, subpartida_counts):
        """
        Guarda un único PDF separado por subpartida.
        """
        try:
            subpartida_base = re.sub(r'[^0-9]', '', subpartida)
            if subpartida_base not in subpartida_counts:
                subpartida_counts[subpartida_base] = 0
            else:
                subpartida_counts[subpartida_base] += 1
            
            if subpartida_counts[subpartida_base] > 0:
                output_filename = f"subpartida_{subpartida_base}_copia_{subpartida_counts[subpartida_base]}.pdf"
            else:
                output_filename = f"subpartida_{subpartida_base}.pdf"
            
            output_path = os.path.join(self.parent_window.output_path.get(), "declaraciones_separadas", output_filename)
            
            with open(output_path, "wb") as output_file:
                writer = PyPDF2.PdfWriter()
                for page_num in pages:
                    writer.add_page(input_pdf.pages[page_num])
                writer.write(output_file)
            
            print(f"Guardado archivo separado: {output_path}")
        except Exception as e:
            print(f"Error al guardar el PDF separado: {str(e)}")

if __name__ == "__main__":
    try:
        root = tk.Tk()
        app = PDFProcessorApp(root)
        root.mainloop()
    except Exception as e:
        print(f"Error al iniciar la aplicación: {str(e)}")