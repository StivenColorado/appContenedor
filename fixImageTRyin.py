import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import pandas as pd
import numpy as np
import os
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image
from datetime import datetime
import logging
import tempfile
import shutil
import io
from PIL import Image as PILImage
from fpdf import FPDF
import re

# Suppress Tkinter deprecation warning
os.environ['TK_SILENCE_DEPRECATION'] = '1'
# Configuración de logging
logging.basicConfig(level=logging.DEBUG)

# Constantes para el tamaño de las imágenes
IMAGE_WIDTH = 90  # Ancho en píxeles
IMAGE_HEIGHT = 90  # Alto en píxeles
EXCEL_START_ROW = 6  # Fila donde comenzarán las imágenes

class ExcelProcessorApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Procesador de Excel")
        self.root.geometry("600x450")
        
        # Configurar el estilo
        style = ttk.Style()
        style.configure('Custom.TFrame', background='#f0f0f0')
        
        # Frame principal
        self.main_frame = ttk.Frame(root, padding="20", style='Custom.TFrame')
        self.main_frame.pack(fill=tk.BOTH, expand=True)
        
        # Etiqueta de título
        self.title_label = ttk.Label(
            self.main_frame,
            text="Procesador de Excel",
            font=('Helvetica', 16, 'bold')
        )
        self.title_label.pack(pady=20)
        
        # Frame para los archivos
        self.file_frame = ttk.Frame(self.main_frame)
        self.file_frame.pack(fill=tk.X, pady=20)
        
        # Input file
        self.input_path = tk.StringVar()
        self.input_label = ttk.Label(
            self.file_frame,
            text="Archivo de entrada:",
            font=('Helvetica', 10)
        )
        self.input_label.pack(anchor=tk.W)
        
        self.input_frame = ttk.Frame(self.file_frame)
        self.input_frame.pack(fill=tk.X, pady=5)
        
        self.input_entry = ttk.Entry(
            self.input_frame,
            textvariable=self.input_path,
            state='readonly'
        )
        self.input_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 5))
        
        self.input_button = ttk.Button(
            self.input_frame,
            text="Seleccionar archivo",
            command=self.select_input_file
        )
        self.input_button.pack(side=tk.RIGHT)
        
        # Número de consolidado
        self.consolidado_frame = ttk.Frame(self.file_frame)
        self.consolidado_frame.pack(fill=tk.X, pady=10)
        
        self.consolidado_label = ttk.Label(
            self.consolidado_frame,
            text="Número de consolidado:",
            font=('Helvetica', 10)
        )
        self.consolidado_label.pack(side=tk.LEFT)
        
        self.consolidado_var = tk.StringVar()
        self.consolidado_entry = ttk.Entry(
            self.consolidado_frame,
            textvariable=self.consolidado_var,
            width=10
        )
        self.consolidado_entry.pack(side=tk.LEFT, padx=5)
        
        # Output directory
        self.output_path = tk.StringVar()
        self.output_label = ttk.Label(
            self.file_frame,
            text="Carpeta de salida:",
            font=('Helvetica', 10)
        )
        self.output_label.pack(anchor=tk.W, pady=(20, 0))
        
        self.output_frame = ttk.Frame(self.file_frame)
        self.output_frame.pack(fill=tk.X, pady=5)
        
        self.output_entry = ttk.Entry(
            self.output_frame,
            textvariable=self.output_path,
            state='readonly'
        )
        self.output_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 5))
        
        self.output_button = ttk.Button(
            self.output_frame,
            text="Seleccionar carpeta",
            command=self.select_output_directory
        )
        self.output_button.pack(side=tk.RIGHT)
        
        # Botón de procesar
        self.process_button = ttk.Button(
            self.main_frame,
            text="Procesar Excel",
            command=self.process_file
        )
        self.process_button.pack(pady=30)
        
        # Status label
        self.status_label = ttk.Label(
            self.main_frame,
            text="",
            font=('Helvetica', 10)
        )
        self.status_label.pack(pady=10)

    def select_input_file(self):
        filename = filedialog.askopenfilename(
            title="Seleccionar archivo Excel",
            filetypes=[("Excel files", "*.xlsx *.xls")]
        )
        if filename:
            self.input_path.set(filename)
            # Intentar extraer el número de consolidado del archivo
            try:
                df = pd.read_excel(filename)
                zafiro_text = None
                for column in df.columns:
                    for value in df[column].astype(str):
                        if 'ZAFIRO-' in value:
                            zafiro_text = value
                            break
                    if zafiro_text:
                        break
                
                if zafiro_text:
                    match = re.search(r'ZAFIRO-(\d+)-\d+', zafiro_text)
                    if match:
                        self.consolidado_var.set(match.group(1))
            except Exception as e:
                logging.error(f"Error extracting consolidado number: {str(e)}")

    def select_output_directory(self):
        directory = filedialog.askdirectory(
            title="Seleccionar carpeta de salida"
        )
        if directory:
            self.output_path.set(directory)

    def process_file(self):
        if not self.input_path.get() or not self.output_path.get() or not self.consolidado_var.get():
            messagebox.showerror("Error", "Por favor complete todos los campos requeridos")
            return
        
        self.status_label.config(text="Procesando archivo...")
        self.process_button.state(['disabled'])
        
        self.root.after(100, self.run_processing)

    def run_processing(self):
        try:
            consolidado = self.consolidado_var.get()
            process_excel(self.input_path.get(), self.output_path.get(), consolidado)
            self.status_label.config(text="¡Archivos procesados correctamente!")
            messagebox.showinfo(
                "Éxito",
                "Los archivos han sido procesados correctamente"
            )
        except Exception as e:
            self.status_label.config(text="Error al procesar los archivos")
            messagebox.showerror("Error", str(e))
        finally:
            self.process_button.state(['!disabled'])

def clean_numeric_value(value):
    if pd.isna(value):
        return 0
    value = str(value).replace('¥', '').replace(' ', '')
    value = value.replace('.', '').replace(',', '.')
    try:
        return float(value)
    except:
        return 0

def find_header_row(df):
    keywords = ['CTNS', 'MARCA', 'CBM', 'WEIGHT', 'PRODUCTO', 'PRODUCT PICTURE']
    for idx, row in df.iterrows():
        row_str = ' '.join(str(val).upper().strip() for val in row)
        matches = sum(keyword in row_str for keyword in keywords)
        if matches >= 3:  # Aumentamos el número mínimo de coincidencias
            return idx
    return None

def find_column(columns, keyword):
    matches = [col for col in columns if keyword in col.upper()]
    if matches:
        return matches[0]
    raise ValueError(f"Columna requerida no encontrada: {keyword}")

def get_color_by_brand(brand):
    predefined_colors = {
        'SONY': 'ADD8E6',
        'XIAOMI': '90EE90',
        'SAMSUNG': 'FFD700',
        'DEFAULT': 'FFFFFF'
    }
    return predefined_colors.get(brand.upper(), predefined_colors['DEFAULT'])

def find_brand_column(df):
    """
    Encuentra la columna que contiene la marca del producto
    """
    possible_names = ['SHIPPING MARK MARCA', 'SHIPPING MARK', 'MARCA']
    for col in df.columns:
        if any(name in str(col).upper() for name in possible_names):
            return col
    return None

def extract_and_save_images_from_workbook(workbook, temp_dir, header_row):
    """
    Extrae y guarda las imágenes con tamaño uniforme manteniendo la información de posición original
    """
    images_info = {
        'header': [],
        'products': {}
    }
    
    for sheet in workbook.worksheets:
        for image in sheet._images:
            img_cell = f"{image.anchor._from.col}_{image.anchor._from.row}"
            img_path = os.path.join(temp_dir, f"image_{img_cell}.png")
            
            try:
                # Extraer y redimensionar la imagen
                image_data = image.ref
                pil_image = PILImage.open(io.BytesIO(image_data.getvalue()))
                
                # Redimensionar la imagen manteniendo la proporción
                pil_image.thumbnail((IMAGE_WIDTH, IMAGE_HEIGHT), PILImage.Resampling.LANCZOS)
                
                # Crear una nueva imagen con fondo transparente
                new_image = PILImage.new('RGBA', (IMAGE_WIDTH, IMAGE_HEIGHT), (255, 255, 255, 0))
                
                # Calcular posición para centrar la imagen
                x = (IMAGE_WIDTH - pil_image.width) // 2
                y = (IMAGE_HEIGHT - pil_image.height) // 2
                
                # Pegar la imagen redimensionada en el centro
                new_image.paste(pil_image, (x, y))
                
                # Guardar la imagen procesada
                new_image.save(img_path, 'PNG')
                
                row_index = image.anchor._from.row
                
                # Guardar información de la imagen con su posición original
                image_info = {
                    'path': img_path,
                    'original_row': row_index,
                    'original_col': image.anchor._from.col
                }
                
                # Clasificar la imagen manteniendo el índice original
                if row_index <= header_row:
                    images_info['header'].append(image_info)
                else:
                    images_info['products'][row_index] = image_info
                
                logging.debug(f"Processed image at row {row_index}, col {image.anchor._from.col}")
                
            except Exception as e:
                logging.error(f"Failed to process image {img_cell}: {str(e)}")
    
    return images_info

def process_brand_excel(brand_df, output_path, marca, year, consolidado, images_info, start_row):
    """
    Procesa el Excel para una marca específica, manejando datos e imágenes simultáneamente
    """
    filename = f"MARCA_{marca}_{consolidado}-{year}.xlsx"
    filepath = os.path.join(output_path, filename)
    
    # Crear un nuevo libro de Excel
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'Datos'
    
    # Configurar encabezados
    for col_idx, column_name in enumerate(brand_df.columns, 1):
        cell = worksheet.cell(row=EXCEL_START_ROW, column=col_idx)
        cell.value = column_name
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Ajustar ancho de columna
        worksheet.column_dimensions[get_column_letter(col_idx)].width = 15
    
    # Encontrar la columna de imágenes de producto
    product_pic_col = None
    for idx, col in enumerate(brand_df.columns):
        if 'PRODUCT PICTURE' in str(col).upper():
            product_pic_col = idx + 1
            worksheet.column_dimensions[get_column_letter(product_pic_col)].width = 15
            break
    
    # Procesar cada fila del DataFrame
    for df_idx, row in brand_df.iterrows():
        excel_row = EXCEL_START_ROW + df_idx + 1  # +1 porque ya escribimos los encabezados
        original_row = start_row + df_idx
        
        # Escribir datos de la fila
        for col_idx, value in enumerate(row, 1):
            cell = worksheet.cell(row=excel_row, column=col_idx)
            cell.value = value
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Ajustar altura de la fila para las imágenes
        worksheet.row_dimensions[excel_row].height = 70
        
        # Agregar imagen si existe
        if product_pic_col and original_row in images_info['products']:
            try:
                img_info = images_info['products'][original_row]
                img = Image(img_info['path'])
                
                # Calcular la celda destino para la imagen
                cell_address = f"{get_column_letter(product_pic_col)}{excel_row}"
                
                # Añadir imagen manteniendo el tamaño uniforme
                worksheet.add_image(img, cell_address)
                logging.debug(f"Added image to {cell_address}")
            except Exception as e:
                logging.error(f"Failed to add image for row {excel_row}: {str(e)}")
    
    # Agregar fórmulas de suma al final
    total_row = EXCEL_START_ROW + len(brand_df) + 1
    
    # Identificar columnas numéricas para las sumas
    numeric_columns = []
    for col_idx, col_name in enumerate(brand_df.columns, 1):
        if any(keyword in str(col_name).upper() for keyword in ['CTNS', 'CBM', 'WEIGHT', 'TOTAL']):
            col_letter = get_column_letter(col_idx)
            start_cell = f"{col_letter}{EXCEL_START_ROW + 1}"
            end_cell = f"{col_letter}{total_row - 1}"
            
            # Agregar fórmula de suma
            sum_cell = worksheet.cell(row=total_row, column=col_idx)
            sum_cell.value = f"=SUM({start_cell}:{end_cell})"
            sum_cell.font = Font(bold=True)
            sum_cell.fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
            
            numeric_columns.append(col_idx)
    
    # Agregar etiqueta "TOTAL" en la primera columna de la fila de totales
    total_label = worksheet.cell(row=total_row, column=1)
    total_label.value = "TOTAL"
    total_label.font = Font(bold=True)
    total_label.fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
    
    # Guardar el archivo
    try:
        workbook.save(filepath)
        logging.debug(f"Successfully saved file: {filepath}")
    except Exception as e:
        logging.error(f"Failed to save file {filepath}: {str(e)}")
        raise
    
def process_excel(input_path, output_path, consolidado):
    try:
        if not input_path.endswith(('.xlsx', '.xls')):
            raise ValueError("El archivo seleccionado no es un archivo de Excel válido.")
        
        # Crear directorio temporal para imágenes
        temp_dir = tempfile.mkdtemp()
        
        # Cargar workbook
        workbook = load_workbook(input_path)
        
        # Encontrar la fila del encabezado
        df_temp = pd.read_excel(input_path, header=None)
        header_row = None
        for idx, row in df_temp.iterrows():
            row_values = [str(val).upper().strip() for val in row.values]
            row_text = ' '.join(row_values)
            if 'SHIPPING MARK MARCA' in row_text or 'PRODUCT PICTURE' in row_text:
                header_row = idx
                break
            
        
        if header_row is None:
            raise ValueError("No se encontró la fila de encabezados")
        
        # Extraer y categorizar imágenes
        images_info = extract_and_save_images_from_workbook(workbook, temp_dir, header_row)
        
        # Leer datos con el encabezado correcto
        df = pd.read_excel(input_path, header=header_row)
        
        # Encontrar la columna de marca
        marca_col = find_brand_column(df)
        if marca_col is None:
            raise ValueError("No se encontró la columna de marca del producto")
        
        # Limpiar nombres de columnas
        df.columns = df.columns.str.strip()
        
        # Procesar por marca
        df[marca_col] = df[marca_col].astype(str).str.strip().str.upper()
        
        for marca in df[marca_col].unique():
            if pd.notna(marca) and marca.strip():
                brand_df = df[df[marca_col] == marca].copy()
                process_brand_excel(brand_df, output_path, marca, 
                                 str(datetime.now().year)[-2:],
                                 consolidado, images_info, header_row + 1)
        
        # Limpiar archivos temporales
        shutil.rmtree(temp_dir)
        logging.debug("Finished processing successfully")
        
    except Exception as e:
        logging.error(f"Error processing file: {str(e)}")
        if 'temp_dir' in locals():
            shutil.rmtree(temp_dir)
        raise
def create_pdf_from_excel(excel_path, pdf_path):
    workbook = load_workbook(excel_path)
    sheet = workbook['RESULTADOS']
    
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Arial", size=10)
    
    for row in sheet.iter_rows(values_only=True):
        row_data = [str(cell) if cell is not None else '' for cell in row]
        pdf.cell(200, 10, txt=" | ".join(row_data).encode('latin-1', 'replace').decode('latin-1'), ln=True)
    
    pdf.output(pdf_path)
def resize_image(image_path, max_width, max_height):
    """
    Redimensiona la imagen para que se ajuste a las dimensiones máximas permitidas
    sin distorsionarla.
    """
    with PILImage.open(image_path) as img:
        # Calcular proporciones y redimensionar
        img.thumbnail((max_width, max_height), PILImage.LANCZOS)
        temp_path = image_path.replace(".png", "_resized.png")
        img.save(temp_path, "PNG")
        return temp_path


def find_real_header_row(df):
    keywords = ['CTNS', 'MARCA', 'CBM', 'WEIGHT', 'PRODUCTO', 'PRODUCT PICTURE']
    for idx, row in df.iterrows():
        row_str = ' '.join(str(val).upper().strip() for val in row)
        matches = sum(keyword in row_str for keyword in keywords)
        if matches >= 3:  # Aumentamos el número mínimo de coincidencias
            return idx
    return None

def copy_workbook_structure(input_path, header_end_row):
    from openpyxl import load_workbook

    wb = load_workbook(input_path)
    ws = wb.active

    images_info = []
    for image in ws._images:
        images_info.append({
            'image': image,
            'row': image.anchor._from.row + 1,
            'col': image.anchor._from.col + 1
        })

    return wb, images_info

def clean_and_validate_brands(df, marca_col):
    # Normalizar datos (strip espacios y manejo de NaN)
    df[marca_col] = df[marca_col].str.strip()
    
    # Eliminar registros con marcas no válidas
    df = df[df[marca_col].notna()]
    
    # Eliminar caracteres especiales innecesarios
    df[marca_col] = df[marca_col].replace(r'[^\w\s]', '', regex=True)
    
    return df
def extract_and_save_images(workbook, temp_dir, header_row):
    """
    Extrae y guarda las imágenes con tamaño uniforme y posicionamiento correcto
    """
    images_info = {
        'header': [],
        'products': {}
    }
    
    # Constante para el offset vertical inicial (5 filas)
    VERTICAL_OFFSET = 5
    
    for sheet in workbook.worksheets:
        for image in sheet._images:
            img_cell = f"{image.anchor._from.col}_{image.anchor._from.row}"
            img_path = os.path.join(temp_dir, f"image_{img_cell}.png")
            
            try:
                # Extraer y redimensionar la imagen
                image_data = image.ref
                pil_image = PILImage.open(io.BytesIO(image_data.getvalue()))
                
                # Redimensionar la imagen manteniendo la proporción
                pil_image.thumbnail((IMAGE_WIDTH, IMAGE_HEIGHT), PILImage.Resampling.LANCZOS)
                
                # Crear una nueva imagen con fondo blanco del tamaño exacto deseado
                new_image = PILImage.new('RGBA', (IMAGE_WIDTH, IMAGE_HEIGHT), (255, 255, 255, 0))
                
                # Calcular posición para centrar la imagen
                x = (IMAGE_WIDTH - pil_image.width) // 2
                y = (IMAGE_HEIGHT - pil_image.height) // 2
                
                # Pegar la imagen redimensionada en el centro
                new_image.paste(pil_image, (x, y))
                
                # Guardar la imagen procesada
                new_image.save(img_path, 'PNG')
                
                # Guardar información de la imagen con posición ajustada
                image_info = {
                    'path': img_path,
                    'row': image.anchor._from.row,
                    'col': image.anchor._from.col,
                    'width': IMAGE_WIDTH,
                    'height': IMAGE_HEIGHT
                }
                
                # Clasificar la imagen y ajustar posición vertical
                if image.anchor._from.row <= header_row:
                    images_info['header'].append(image_info)
                else:
                    # Ajustar la posición vertical para las imágenes de productos
                    adjusted_row = VERTICAL_OFFSET + (image.anchor._from.row - header_row)
                    images_info['products'][adjusted_row] = image_info
                
                logging.debug(f"Processed image at adjusted row {adjusted_row if 'adjusted_row' in locals() else image.anchor._from.row}, col {image.anchor._from.col}")
            
            except Exception as e:
                logging.error(f"Failed to process image {img_cell}: {str(e)}")
    
    return images_info

def main():
    root = tk.Tk()
    app = ExcelProcessorApp(root)
    root.mainloop()

if __name__ == "__main__":
    main()