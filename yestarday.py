# -*- coding: utf-8 -*-
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import pandas as pd
import numpy as np
import os
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image
from datetime import datetime
import io
from PIL import Image as PILImage
import openpyxl

class ExcelProcessorApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Procesador de Excel")
        self.root.geometry("600x450")
        # Agregar icono a la ventana
        try:
            self.root.iconbitmap('icon.ico')  # Asegúrate de tener el archivo icon.ico en el mismo directorio
        except:
            pass
        
        # Configurar el estilo correctamente
        style = ttk.Style(root)
        style.configure('TFrame', background='#f0f0f0')  # Usa 'TFrame' directamente para verificar si el problema persiste

        # Frame principal
        self.main_frame = ttk.Frame(root, padding="20", style='Custom.TFrame')
        self.main_frame.pack(fill=tk.BOTH, expand=True)
        
        # Etiqueta de titulo
        self.title_label = ttk.Label(
            self.main_frame,
            text="Procesador de Excel",
            font=('Helvetica', 16, 'bold')
        )
        self.title_label.pack(pady=20)
        
        # Frame para los archivos y numero de consolidado
        self.file_frame = ttk.Frame(self.main_frame)
        self.file_frame.pack(fill=tk.X, pady=20)
        
        # Numero de consolidado
        self.consolidado_frame = ttk.Frame(self.file_frame)
        self.consolidado_frame.pack(fill=tk.X, pady=5)
        
        self.consolidado_label = ttk.Label(
            self.consolidado_frame,
            text="Numero de Consolidado:",
            font=('Helvetica', 10)
        )
        self.consolidado_label.pack(side=tk.LEFT)
        
        self.consolidado_var = tk.StringVar()
        self.consolidado_entry = ttk.Entry(
            self.consolidado_frame,
            textvariable=self.consolidado_var,
            width=10
        )
        self.consolidado_entry.pack(side=tk.LEFT, padx=5)
        
        # Input file
        self.input_path = tk.StringVar()
        self.input_label = ttk.Label(
            self.file_frame,
            text="Archivo de entrada:",
            font=('Helvetica', 10)
        )
        self.input_label.pack(anchor=tk.W)
        
        self.input_frame = ttk.Frame(self.file_frame)
        self.input_frame.pack(fill=tk.X, pady=5)
        
        self.input_entry = ttk.Entry(
            self.input_frame,
            textvariable=self.input_path,
            state='readonly'
        )
        self.input_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 5))
        
        self.input_button = ttk.Button(
            self.input_frame,
            text="Seleccionar archivo",
            command=self.select_input_file
        )
        self.input_button.pack(side=tk.RIGHT)
        
        # Output directory
        self.output_path = tk.StringVar()
        self.output_label = ttk.Label(
            self.file_frame,
            text="Carpeta de salida:",
            font=('Helvetica', 10)
        )
        self.output_label.pack(anchor=tk.W, pady=(20, 0))
        
        self.output_frame = ttk.Frame(self.file_frame)
        self.output_frame.pack(fill=tk.X, pady=5)
        
        self.output_entry = ttk.Entry(
            self.output_frame,
            textvariable=self.output_path,
            state='readonly'
        )
        self.output_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 5))
        
        self.output_button = ttk.Button(
            self.output_frame,
            text="Seleccionar carpeta",
            command=self.select_output_directory
        )
        self.output_button.pack(side=tk.RIGHT)
        
        # Boton de procesar
        self.process_button = ttk.Button(
            self.main_frame,
            text="Procesar Excel",
            command=self.process_file
        )
        self.process_button.pack(pady=30)
        
        # Barra de progreso
        self.progress = ttk.Progressbar(
            self.main_frame,
            orient=tk.HORIZONTAL,
            length=300,
            mode='indeterminate'
        )
        self.progress.pack(pady=10)
        
        # Status label
        self.status_label = ttk.Label(
            self.main_frame,
            text="",
            font=('Helvetica', 10)
        )
        self.status_label.pack(pady=10)

    def select_input_file(self):
        filename = filedialog.askopenfilename(
            title="Seleccionar archivo Excel",
            filetypes=[("Excel files", "*.xlsx *.xls")]
        )
        if filename:
            self.input_path.set(filename)

    def select_output_directory(self):
        directory = filedialog.askdirectory(
            title="Seleccionar carpeta de salida"
        )
        if directory:
            self.output_path.set(directory)

    def process_file(self):
        if not self.input_path.get() or not self.output_path.get() or not self.consolidado_var.get():
            messagebox.showerror(
                "Error",
                "Por favor complete todos los campos requeridos"
            )
            return
        
        try:
            consolidado = int(self.consolidado_var.get())
        except ValueError:
            messagebox.showerror(
                "Error",
                "El numero de consolidado debe ser un numero valido"
            )
            return
        
        self.progress.start()
        self.status_label.config(text="Procesando archivo...")
        self.process_button.state(['disabled'])
        
        self.root.after(100, lambda: self.run_processing(consolidado))

    def run_processing(self, consolidado):
        try:
            process_excel(self.input_path.get(), self.output_path.get(), consolidado)
            self.progress.stop()
            self.status_label.config(text="¡Archivos procesados correctamente!")
            messagebox.showinfo(
                "Exito",
                "Los archivos han sido procesados correctamente"
            )
        except Exception as e:
            self.progress.stop()
            self.status_label.config(text="Error al procesar los archivos")
            messagebox.showerror("Error", str(e))
        finally:
            self.process_button.state(['!disabled'])
            self.progress.stop()

def get_header_rows(input_path):
    """Obtiene las primeras 7 filas del archivo original"""
    try:
        header_df = pd.read_excel(input_path, nrows=7, header=None)
        return header_df
    except Exception as e:
        print(f"Error al leer las filas de encabezado: {str(e)}")
        return None

def find_header_row(df, start_row=7):
    """Busca la fila de encabezados después de las 7 primeras filas"""
    keywords = ['CTNS', 'MARCA', 'CBM', 'WEIGHT', 'PRODUCTO', 'PRODUCT PICTURE']
    for idx, row in df.iloc[start_row:].iterrows():
        row_str = ' '.join(str(val).upper().strip() for val in row)
        matches = sum(keyword in row_str for keyword in keywords)
        if matches >= 3:
            return idx
    return None

def get_color_by_brand(brand):
    predefined_colors = {
        'SONY': 'ADD8E6',
        'XIAOMI': '90EE90',
        'SAMSUNG': 'FFD700',
        'DEFAULT': 'FFFFFF'
    }
    return predefined_colors.get(brand.upper(), predefined_colors['DEFAULT'])

def find_real_header_row(df):
    """
    Busca la fila que contiene los encabezados específicos mencionados
    """
    header_keywords = [
        'ESPECIAL O NORMAL',
        'FECHA DE ENTREGA',
        'SHIPPER',
        'SHIPPING MARK',
        'PRODUCT PICTURE',
        'MARCA',
        '品牌',  # Agregamos el término en chino
        'PRODUCT DESCRIPTION',
        'CTNS',
        'QTY/CTN'
    ]
    
    for idx, row in df.iterrows():
        row_str = ' '.join(str(val).upper().strip() for val in row if pd.notna(val))
        matches = sum(keyword in row_str.upper() for keyword in header_keywords)
        if matches >= 4:  # Reducimos el número de coincidencias necesarias
            return idx
    return None

def copy_workbook_structure(input_path, header_end_row):
    """
    Copia la estructura exacta del archivo original, incluyendo imágenes y formato
    """
    src_wb = openpyxl.load_workbook(input_path)
    src_ws = src_wb.active
    
    # Guardar información de imágenes y sus posiciones
    images_info = []
    for image in src_ws._images:
        img_cell = image.anchor._from
        row_idx = img_cell.row
        col_idx = img_cell.col
        if row_idx <= header_end_row:
            images_info.append({
                'image': image,
                'row': row_idx,
                'col': col_idx
            })
    
    return src_wb, images_info

def process_brand_excel(brand_df, input_path, header_end_row, output_path, marca, year, consolidado):
    filename = f"MARCA {marca} {consolidado}-{year}.xlsx"
    filepath = os.path.join(output_path, filename)
    
    # Copiar estructura original incluyendo imágenes
    src_wb, images_info = copy_workbook_structure(input_path, header_end_row)
    src_ws = src_wb.active
    
    # Crear nuevo archivo
    with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
        # Primero, copiar las filas de encabezado del archivo original
        header_df = pd.read_excel(input_path, nrows=header_end_row+1, header=None)
        header_df.to_excel(writer, sheet_name='Datos', index=False, header=False)
        
        # Luego escribir los datos de la marca
        brand_df = brand_df.drop(columns=['UNIT PRICE (RMB)', 'AMOUNT (RMB)'], errors='ignore')
        brand_df.to_excel(writer, sheet_name='Datos', startrow=header_end_row+1, index=False)
        
        workbook = writer.book
        worksheet = writer.sheets['Datos']
        
        # Copiar imágenes del encabezado
        for img_info in images_info:
            try:
                img_path = img_info['image'].path  # Ruta de la imagen original
                if os.path.exists(img_path):
                    img = Image(img_path)
                    cell = worksheet.cell(row=img_info['row'], column=img_info['col'])
                    worksheet.add_image(img, cell.coordinate)
            except Exception as e:
                print(f"Error al copiar imagen: {str(e)}")
        
        # Copiar formato y estilos de las celdas del encabezado
        for row in range(1, header_end_row + 2):
            for col in range(1, src_ws.max_column + 1):
                src_cell = src_ws.cell(row=row, column=col)
                dst_cell = worksheet.cell(row=row, column=col)
                
                # Copiar formato
                if src_cell.font:
                    dst_cell.font = Font(
                        name=src_cell.font.name,
                        size=src_cell.font.size,
                        bold=src_cell.font.bold,
                        italic=src_cell.font.italic,
                        vertAlign=src_cell.font.vertAlign,
                        underline=src_cell.font.underline,
                        strike=src_cell.font.strike,
                        color=src_cell.font.color
                    )
                if src_cell.fill:
                    dst_cell.fill = PatternFill(
                        fill_type=src_cell.fill.fill_type,
                        start_color=src_cell.fill.start_color,
                        end_color=src_cell.fill.end_color
                    )
                if src_cell.alignment:
                    dst_cell.alignment = Alignment(
                        horizontal=src_cell.alignment.horizontal,
                        vertical=src_cell.alignment.vertical,
                        text_rotation=src_cell.alignment.text_rotation,
                        wrap_text=src_cell.alignment.wrap_text,
                        shrink_to_fit=src_cell.alignment.shrink_to_fit,
                        indent=src_cell.alignment.indent
                    )
        
        # Aplicar color de marca a las filas de datos
        color_hex = get_color_by_brand(marca)
        fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type="solid")
        for row in range(header_end_row + 2, header_end_row + len(brand_df) + 2):
            for col in range(1, len(brand_df.columns) + 1):
                worksheet.cell(row=row, column=col).fill = fill
        
        # Procesar sumas para columnas totales
        total_columns = {
            'T/CBM': 'T/CBM',
            'T/WEIGHT (KG)': 'T/WEIGHT (KG)',
            'CTNS': 'CTNS'
        }
        
        for display_name, col_name in total_columns.items():
            if col_name in brand_df.columns:
                col_index = list(brand_df.columns).index(col_name) + 1
                column_letter = get_column_letter(col_index)
                data_start_row = header_end_row + 2
                data_end_row = header_end_row + len(brand_df) + 1
                formula = f'=SUM({column_letter}{data_start_row}:{column_letter}{data_end_row})'
                sum_cell = f'{column_letter}{data_end_row + 1}'
                worksheet[sum_cell] = formula
                worksheet[sum_cell].font = Font(bold=True)
                worksheet[sum_cell].alignment = Alignment(horizontal='right')

def process_excel(input_path, output_path, consolidado):
    try:
        # Leer el archivo completo primero sin especificar encabezados
        df_raw = pd.read_excel(input_path, header=None)
        
        # Encontrar la fila real de encabezados
        header_row = find_real_header_row(df_raw)
        if header_row is None:
            raise ValueError("No se encontró la fila de encabezados adecuada")
        
        # Leer el archivo nuevamente usando la fila de encabezados correcta
        df = pd.read_excel(input_path, header=header_row)

        # Eliminar columnas UNIT PRICE (RMB) y AMOUNT (RMB) si existen
        df = df.drop(columns=['UNIT PRICE (RMB)', 'AMOUNT (RMB)'], errors='ignore')
        
        # Identificar columna de marca
        marca_col = None
        for col in df.columns:
            if any(keyword in str(col).upper() for keyword in ['品牌', 'MARCA']):
                marca_col = col
                break
                
        if marca_col is None:
            raise ValueError("No se encontró la columna de marca")
        
        # Obtener año actual
        current_year = str(datetime.now().year)[-2:]
        
        # Limpiar y convertir la columna de marca a string
        df[marca_col] = df[marca_col].fillna('SIN MARCA').astype(str)
        
        # Procesar cada marca por separado
        unique_brands = df[marca_col].unique()
        for marca in unique_brands:
            if marca and marca.strip():  # Verificar que la marca no esté vacía
                brand_df = df[df[marca_col] == marca].copy()
                process_brand_excel(brand_df, input_path, header_row, output_path, marca, current_year, consolidado)
        
        # Crear archivo de resumen general
        summary_filename = f"RESUMEN_GENERAL_CONSO_{consolidado}-{current_year}.xlsx"
        summary_filepath = os.path.join(output_path, summary_filename)
        
        with pd.ExcelWriter(summary_filepath, engine='openpyxl') as writer:
            # Copiar estructura de encabezado
            header_df = pd.read_excel(input_path, nrows=header_row+1, header=None)
            header_df.to_excel(writer, sheet_name='Datos', index=False, header=False)
            
            # Escribir datos principales
            df.to_excel(writer, sheet_name='Datos', startrow=header_row+1, index=False)
            
            # Crear hoja RESULTADOS
            try:
                summary_columns = {
                    'CTNS': 'CARTONES',
                    'T/CBM': 'CUBICAJE',
                    'T/WEIGHT (KG)': 'PESO'
                }
                
                # Buscar las columnas correctas en el DataFrame
                agg_dict = {}
                for old_col in summary_columns.keys():
                    matching_cols = [col for col in df.columns if old_col in str(col)]
                    if matching_cols:
                        agg_dict[matching_cols[0]] = 'sum'
                
                if agg_dict:
                    summary = df.groupby(marca_col).agg(agg_dict).reset_index()
                    # Renombrar columnas
                    new_columns = [marca_col]
                    for col in summary.columns[1:]:
                        for old_col, new_col in summary_columns.items():
                            if old_col in str(col):
                                new_columns.append(new_col)
                                break
                        else:
                            new_columns.append(col)
                    summary.columns = new_columns
                    
                    summary.to_excel(writer, sheet_name='RESULTADOS', index=False)
                    
                    # Asegurar que la hoja RESULTADOS esté visible
                    writer.book['RESULTADOS'].sheet_state = 'visible'
                    
                    # Agregar fila de totales
                    workbook = writer.book
                    worksheet = writer.sheets['RESULTADOS']
                    row_num = len(summary) + 2
                    worksheet.cell(row=row_num, column=1, value='TOTAL')
                    
                    for col_idx, col in enumerate(summary.columns[1:], start=2):
                        total_value = summary[col].sum()
                        worksheet.cell(row=row_num, column=col_idx, value=total_value)
                        worksheet.cell(row=row_num, column=col_idx).font = Font(bold=True)
                        worksheet.cell(row=row_num, column=col_idx).alignment = Alignment(horizontal='right')
            except Exception as e:
                print(f"Error al crear la hoja de resultados: {str(e)}")
            
            # Asegurar que la hoja Datos esté visible
            writer.book['Datos'].sheet_state = 'visible'
    except Exception as e:
        print(f"Error al procesar el archivo: {str(e)}")
        
def main():
    root = tk.Tk()
    app = ExcelProcessorApp(root)
    root.mainloop()

if __name__ == "__main__":
    main()