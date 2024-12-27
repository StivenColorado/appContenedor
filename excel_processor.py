import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import pandas as pd
import numpy as np
import os
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image
from datetime import datetime
import logging
import tempfile
import shutil
import io
from PIL import Image as PILImage
from fpdf import FPDF
import re
import time
import copy
from PIL import JpegImagePlugin
from tkinter import simpledialog

JpegImagePlugin._getmp = lambda: None
# Suppress Tkinter deprecation warning
os.environ['TK_SILENCE_DEPRECATION'] = '1'
# Configuración de logging
logging.basicConfig(level=logging.DEBUG)

# Constantes para el tamaño de las imágenes
IMAGE_WIDTH = 90  # Ancho en píxeles
IMAGE_HEIGHT = 90  # Alto en píxeles
HEADER_IMAGE_WIDTH = 200  # Ancho en píxeles para imágenes de cabecera
HEADER_IMAGE_HEIGHT = 200  # Alto en píxeles para imágenes de cabecera
EXCEL_START_ROW = 6  # Fila donde comenzarán las imágenes

class LoadingScreen:
    def __init__(self):
        self.window = tk.Toplevel()
        self.window.title("")
        self.window.overrideredirect(True)  # Elimina los bordes de la ventana
        
        # Obtener dimensiones de la pantalla
        screen_width = self.window.winfo_screenwidth()
        screen_height = self.window.winfo_screenheight()
        
        # Calcular posición para centrar
        width = 300
        height = 150
        x = (screen_width - width) // 2
        y = (screen_height - height) // 2
        
        self.window.geometry(f"{width}x{height}+{x}+{y}")
        self.window.configure(bg='#f0f0f0')
        
        # Marco principal
        main_frame = ttk.Frame(self.window, style='Custom.TFrame')
        main_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)
        
        # Etiqueta de carga
        self.loading_label = ttk.Label(
            main_frame,
            text="Cargando...",
            font=('Helvetica', 12)
        )
        self.loading_label.pack(pady=10)
        
        # Barra de progreso
        self.progress = ttk.Progressbar(
            main_frame,
            length=200,
            mode='determinate'
        )
        self.progress.pack(pady=10)
        
        # Etiqueta de desarrollador
        self.dev_label = ttk.Label(
            main_frame,
            text="Desarrollado por Stiven Colorado",
            font=('Helvetica', 10, 'italic'),
            foreground='black'
        )
        self.dev_label.pack(pady=10)
        
    def update_progress(self, value):
        self.progress['value'] = value
        self.window.update()
    
    def close(self):
        self.window.destroy()

def center_window(window):
    window.update_idletasks()
    width = window.winfo_width()
    height = window.winfo_height()
    x = (window.winfo_screenwidth() // 2) - (width // 2)
    y = (window.winfo_screenheight() // 2) - (height // 2)
    window.geometry(f'{width}x{height}+{x}+{y}')



def resource_path(relative_path):
    """ Get absolute path to resource, works for dev and for PyInstaller """
    try:
        # PyInstaller creates a temp folder and stores path in _MEIPASS
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")

    return os.path.join(base_path, relative_path)# funciones para la ventana principal

def setup_app_icon(root):
    """
    Configura el icono de la aplicación solo para la barra de tareas de Windows
    """
    try:
        if os.name == 'nt':  # Solo para Windows
            icon_path = resource_path('icon.ico')
            root.iconbitmap(icon_path)
        
        # Quitar el icono de la ventana
        root.overrideredirect(False)
        root.iconwindow()
        
    except Exception as e:
        logging.error(f"Error setting up app icon: {str(e)}")
class ExcelProcessorApp:
    def __init__(self, root):
        self.root = root
        
        setup_app_icon(root)
        
        # Ocultar la ventana principal temporalmente
        self.root.withdraw()
        
        # Mostrar pantalla de carga
        loading = LoadingScreen()
        for i in range(0, 101, 2):
            loading.update_progress(i)
            time.sleep(0.06)  # Total ~3 segundos
        loading.close()
        # mostrar la pantalla
        self.root.deiconify()
        self.root.title("Procesador de Excel")
        self.root.geometry("700x500")
        center_window(root)
        
        # Configurar el estilo
        style = ttk.Style()
        style.configure('Custom.TFrame', background='#f0f0f0')
        
        # Frame principal
        self.main_frame = ttk.Frame(root, padding="20", style='Custom.TFrame')
        self.main_frame.pack(fill=tk.BOTH, expand=True)
        
        # Logo Zafiro
        try:
            zafiro_path = resource_path('zafiro.png')
            zafiro_image = tk.PhotoImage(file=zafiro_path)
            # Redimensionar si es necesario
            zafiro_image = zafiro_image.subsample(2, 2)  # Ajusta estos números según necesites
            zafiro_label = ttk.Label(self.main_frame, image=zafiro_image)
            zafiro_label.image = zafiro_image  # Mantener referencia
            zafiro_label.pack(pady=(0, 20))
        except Exception as e:
            logging.error(f"Error loading Zafiro logo: {str(e)}")

        # Etiqueta de título
        self.title_label = ttk.Label(
            self.main_frame,
            text="Procesador de Excel",
            font=('Helvetica', 16, 'bold')
        )
        self.title_label.pack(pady=20)
        
        # Frame para los archivos
        self.file_frame = ttk.Frame(self.main_frame)
        self.file_frame.pack(fill=tk.X, pady=20)
        
        # Input file
        self.input_path = tk.StringVar()
        self.input_label = ttk.Label(
            self.file_frame,
            text="Archivo de entrada:",
            font=('Helvetica', 10)
        )
        self.input_label.pack(anchor=tk.W)
        
        self.input_frame = ttk.Frame(self.file_frame)
        self.input_frame.pack(fill=tk.X, pady=5)
        
        self.input_entry = ttk.Entry(
            self.input_frame,
            textvariable=self.input_path,
            state='readonly'
        )
        self.input_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 5))
        
        self.input_button = ttk.Button(
            self.input_frame,
            text="Seleccionar archivo",
            command=self.select_input_file
        )
        self.input_button.pack(side=tk.RIGHT)
        
        # Número de consolidado
        self.consolidado_frame = ttk.Frame(self.file_frame)
        self.consolidado_frame.pack(fill=tk.X, pady=10)
        
        self.consolidado_label = ttk.Label(
            self.consolidado_frame,
            text="Número de consolidado:",
            font=('Helvetica', 10)
        )
        self.consolidado_label.pack(side=tk.LEFT)
        
        self.consolidado_var = tk.StringVar()
        self.consolidado_entry = ttk.Entry(
            self.consolidado_frame,
            textvariable=self.consolidado_var,
            width=10
        )
        self.consolidado_entry.pack(side=tk.LEFT, padx=5)
        
        # Output directory
        self.output_path = tk.StringVar()
        self.output_label = ttk.Label(
            self.file_frame,
            text="Carpeta de salida:",
            font=('Helvetica', 10)
        )
        self.output_label.pack(anchor=tk.W, pady=(20, 0))
        
        self.output_frame = ttk.Frame(self.file_frame)
        self.output_frame.pack(fill=tk.X, pady=5)
        
        self.output_entry = ttk.Entry(
            self.output_frame,
            textvariable=self.output_path,
            state='readonly'
        )
        self.output_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 5))
        
        self.output_button = ttk.Button(
            self.output_frame,
            text="Seleccionar carpeta",
            command=self.select_output_directory
        )
        self.output_button.pack(side=tk.RIGHT)
        
        # Botón de procesar
        self.process_button = ttk.Button(
            self.main_frame,
            text="Procesar Excel",
            command=self.process_file
        )
        self.process_button.pack(pady=30)
        
        # Status label
        self.status_label = ttk.Label(
            self.main_frame,
            text="",
            font=('Helvetica', 10)
        )
        self.status_label.pack(pady=10)

    def validate_required_columns(self, df):
        required_columns = ['CTNS', 'T/CBM', 'T/WEIGHT (KG)']
        missing_columns = []
        
        for col in required_columns:
            found = False
            for df_col in df.columns:
                if col in str(df_col).upper():
                    found = True
                    break
            if not found:
                missing_columns.append(col)
        
        if missing_columns:
            filename = os.path.basename(self.input_path.get())
            error_msg = f"El archivo '{filename}' no cuenta con las siguientes columnas necesarias para funcionar correctamente:\n\n"
            error_msg += "\n".join(f"- {col}" for col in missing_columns)
            error_msg += "\n\nEsto hará imposible mostrar los resultados esperados."
            messagebox.showerror("Error - Columnas faltantes", error_msg)
            return False
        return True
    
    def select_input_file(self):
        filename = filedialog.askopenfilename(
            title="Seleccionar archivo Excel",
            filetypes=[("Excel files", "*.xlsx *.xls")]
        )
        if filename:
            self.input_path.set(filename)
            # Intentar extraer el número de consolidado del archivo
            try:
                df = pd.read_excel(filename)
                zafiro_text = None
                for column in df.columns:
                    for value in df[column].astype(str):
                        if 'ZAFIRO-' in value:
                            zafiro_text = value
                            break
                    if zafiro_text:
                        break
                
                if zafiro_text:
                    match = re.search(r'ZAFIRO-(\d+)-\d+', zafiro_text)
                    if match:
                        self.consolidado_var.set(match.group(1))
            except Exception as e:
                logging.error(f"Error extracting consolidado number: {str(e)}")

    def select_output_directory(self):
        directory = filedialog.askdirectory(
            title="Seleccionar carpeta de salida"
        )
        if directory:
            self.output_path.set(directory)

    def process_file(self):
        if not self.input_path.get() or not self.output_path.get() or not self.consolidado_var.get():
            messagebox.showerror("Error", "Por favor complete todos los campos requeridos")
            return
        
        self.status_label.config(text="Procesando archivo...")
        self.process_button.state(['disabled'])
        
        self.root.after(100, self.run_processing)

    def run_processing(self):
        try:
            consolidado = self.consolidado_var.get()
            process_excel(self.input_path.get(), self.output_path.get(), consolidado)
            self.status_label.config(text="¡Archivos procesados correctamente!")
            messagebox.showinfo(
                "Éxito",
                "Los archivos han sido procesados correctamente"
            )
        except Exception as e:
            self.status_label.config(text="Error al procesar los archivos")
            messagebox.showerror("Error", str(e))
        finally:
            self.process_button.state(['!disabled'])

def clean_numeric_value(value):
    if pd.isna(value):
        return 0
    value = str(value).replace('¥', '').replace(' ', '')
    value = value.replace('.', '').replace(',', '.')
    try:
        return float(value)
    except:
        return 0

def find_header_row(df):
    keywords = ['CTNS', 'MARCA', 'CBM', 'WEIGHT', 'PRODUCTO', 'PRODUCT PICTURE']
    for idx, row in df.iterrows():
        row_str = ' '.join(str(val).upper().strip() for val in row)
        matches = sum(keyword in row_str for keyword in keywords)
        if matches >= 3:  # Aumentamos el número mínimo de coincidencias
            return idx
    return None

def find_column(columns, keyword):
    matches = [col for col in columns if keyword in col.upper()]
    if matches:
        return matches[0]
    raise ValueError(f"Columna requerida no encontrada: {keyword}")

def get_color_by_brand(brand):
    predefined_colors = {
        'SONY': 'ADD8E6',
        'XIAOMI': '90EE90',
        'SAMSUNG': 'FFD700',
        'DEFAULT': 'FFFFFF'
    }
    return predefined_colors.get(brand.upper(), predefined_colors['DEFAULT'])

def find_brand_column(df):
    """
    Encuentra la columna que contiene la marca del producto
    """
    possible_names = ['SHIPPING MARK MARCA', 'SHIPPING MARK', 'MARCA']
    for col in df.columns:
        if any(name in str(col).upper() for name in possible_names):
            return col
    return None

def extract_and_save_images_from_workbook(workbook, temp_dir, header_row):
    images_info = {
        'header': [],
        'products': {}
    }
    
    HEADER_IMAGE_WIDTH = 400
    HEADER_IMAGE_HEIGHT = 400
    PRODUCT_IMAGE_WIDTH = 90
    PRODUCT_IMAGE_HEIGHT = 90
    
    for sheet in workbook.worksheets:
        for image in sheet._images:
            img_cell = f"{image.anchor._from.col}_{image.anchor._from.row}"
            img_path = os.path.join(temp_dir, f"image_{img_cell}.png")
            
            try:
                image_data = image.ref
                pil_image = PILImage.open(io.BytesIO(image_data.getvalue()))
                
                target_width = HEADER_IMAGE_WIDTH if image.anchor._from.row <= header_row else PRODUCT_IMAGE_WIDTH
                target_height = HEADER_IMAGE_HEIGHT if image.anchor._from.row <= header_row else PRODUCT_IMAGE_HEIGHT
                
                pil_image.thumbnail((target_width, target_height), PILImage.Resampling.LANCZOS)
                new_image = PILImage.new('RGBA', (target_width, target_height), (255, 255, 255, 0))
                
                x = (target_width - pil_image.width) // 2
                y = (target_height - pil_image.height) // 2
                new_image.paste(pil_image, (x, y))
                new_image.save(img_path, 'PNG')
                
                row_index = image.anchor._from.row
                image_info = {
                    'path': img_path,
                    'original_row': row_index,
                    'original_col': image.anchor._from.col
                }
                
                if row_index <= header_row:
                    images_info['header'].append(image_info)
                else:
                    images_info['products'][row_index] = image_info
                
            except Exception as e:
                logging.error(f"Failed to process image {img_cell}: {str(e)}")
    
    return images_info
def process_brand_excel(brand_df, output_path, marca, year, consolidado, images_info, start_row, zafiro_number):
    if 'ANOTACION' in brand_df.columns:
        brand_df = brand_df.drop('ANOTACION', axis=1)
    filename = f"MARCA_{marca}_CONSO_{consolidado}-{year}.xlsx"
    filepath = os.path.join(output_path, filename)
    
    wb = Workbook()
    ws = wb.active
    ws.title = 'Datos'
    
    # Set header row heights
    for i in range(1, 6):
        ws.row_dimensions[i].height = 15  # Reducir el alto de las filas de la 1 a la 5
    
    # Add header images
    header_positions = [
        # y, x, height, width
        (1, 19, 6.69, 2.50),  # Esquina superior izquierda, tercera imagen 
        (1, 17, 4.56, 2.50),  # Parte superior derecha, segunda imagen
        (1, 1, 6.91, 2.70),   # Parte superior derecha, primera imagen
        (1, 22, 4.56, 2.50),  # Parte superior derecha, tamaño reducido, cuarta imagen
    ]
    for idx, header_img in enumerate(images_info['header']):
        try:
            img = Image(header_img['path'])
            img.width = header_positions[idx][2] * 37.795275591  # Convert cm to pixels
            img.height = header_positions[idx][3] * 37.795275591  # Convert cm to pixels
            # Set the position of the image
            cell = ws.cell(row=header_positions[idx][0], column=header_positions[idx][1])
            img.anchor = cell.coordinate
            ws.add_image(img)
        except Exception as e:
            logging.error(f"Failed to process image {header_img['path']}: {str(e)}")
    
    # Add header texts
    ws.merge_cells('D2:E3')
    ws.cell(row=2, column=4).value = "PACKING LIST"
    ws.cell(row=2, column=4).font = Font(bold=True, size=14)
    ws.cell(row=2, column=4).alignment = Alignment(horizontal='center')

    ws.merge_cells('G2:H3')
    ws.cell(row=2, column=7).value = zafiro_number
    ws.cell(row=2, column=7).font = Font(bold=True, size=14, color="FF0000")
    ws.cell(row=2, column=7).alignment = Alignment(horizontal='center')
    
    # Write column headers
    for col_idx, column_name in enumerate(brand_df.columns, 1):
        cell = ws.cell(row=6, column=col_idx)
        cell.value = column_name
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
        ws.column_dimensions[get_column_letter(col_idx)].width = 15
        cell.fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")  # Color azul
    
    # Write data
    current_row = 7
    for df_idx, row in brand_df.iterrows():
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=current_row, column=col_idx)
            cell.value = value
            cell.alignment = Alignment(horizontal='center')
        
        ws.row_dimensions[current_row].height = 70
        
        if start_row + df_idx in images_info['products']:
            product_pic_col = next((i for i, col in enumerate(brand_df.columns, 1) 
                                  if 'PRODUCT PICTURE' in str(col).upper()), None)
            if product_pic_col:
                try:
                    img = Image(images_info['products'][start_row + df_idx]['path'])
                    img.width = 120  # Ajusta el ancho de la imagen
                    img.height = 120  # Ajusta el alto de la imagen
                    cell_address = f"{get_column_letter(product_pic_col)}{current_row}"
                    ws.column_dimensions[get_column_letter(product_pic_col)].width = 25  # Ajusta el ancho de la columna
                    ws.row_dimensions[current_row].height = 100  # Ajusta el alto de la fila
                    ws.add_image(img, cell_address)
                except Exception as e:
                    logging.error(f"Failed to add product image for row {current_row}: {str(e)}")
        
        current_row += 1
    
    # Add totals for specific columns only
    total_row = current_row
    total_columns = ['CTNS', 'T/CBM', 'T/WEIGHT (KG)']
    
    ws.cell(row=total_row, column=1).value = "TOTAL"
    ws.cell(row=total_row, column=1).font = Font(bold=True)
    ws.cell(row=total_row, column=1).fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
    
    for col_name in total_columns:
        if col_name in brand_df.columns:
            col_idx = list(brand_df.columns).index(col_name) + 1
            col_letter = get_column_letter(col_idx)
            sum_cell = ws.cell(row=total_row, column=col_idx)
            sum_cell.value = f"=SUM({col_letter}7:{col_letter}{total_row-1})"
            sum_cell.font = Font(bold=True)
            sum_cell.fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
    
    wb.save(filepath)
def add_totals_row(worksheet, brand_df, total_row):
    """Helper function to add totals row"""
    for col_idx, col_name in enumerate(brand_df.columns, 1):
        if any(keyword in str(col_name).upper() for keyword in ['CTNS', 'CBM', 'WEIGHT', 'TOTAL']):
            col_letter = get_column_letter(col_idx)
            start_cell = f"{col_letter}7"  # Comenzar desde fila 7
            end_cell = f"{col_letter}{total_row - 1}"
            
            sum_cell = worksheet.cell(row=total_row, column=col_idx)
            sum_cell.value = f"=SUM({start_cell}:{end_cell})"
            sum_cell.font = Font(bold=True)
            sum_cell.fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
    
    total_label = worksheet.cell(row=total_row, column=1)
    total_label.value = "TOTAL"
    total_label.font = Font(bold=True)
    total_label.fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")

def find_end_row(df):
    """
    Encuentra la última fila válida de datos antes de la palabra TOTAL
    """
    for idx, row in df.iterrows():
        row_values = [str(val).strip().upper() for val in row if pd.notna(val)]
        if 'TOTAL' in row_values:
            return idx
    return len(df)

def clean_dataframe(df):
    """
    Limpia el DataFrame eliminando filas después de TOTAL y filas que no son datos válidos
    """
    # Encontrar la última fila válida
    end_row = find_end_row(df)
    
    # Cortar el DataFrame hasta la fila encontrada
    df = df.iloc[:end_row]
    
    # Eliminar filas donde todas las columnas numéricas son NaN o 0
    numeric_columns = df.select_dtypes(include=[np.number]).columns
    if not numeric_columns.empty:
        df = df.dropna(subset=numeric_columns, how='all')
        df = df[(df[numeric_columns] != 0).any(axis=1)]
    
    return df

def normalize_brand(brand):
    """
    Normaliza el nombre de la marca eliminando todo después del primer guion (- o _).
    """
    if '-' in brand:
        brand = brand.split('-')[0].strip()
    elif '_' in brand:
        brand = brand.split('_')[0].strip()
    return brand

def normalize_brand_name(brand_name):
    """
    Normaliza el nombre de la marca eliminando caracteres especiales y espacios.
    """
    return re.sub(r'[^A-Za-z0-9]', '', brand_name).upper()

def validate_and_normalize_brands(df, marca_col):
    """
    Normaliza y valida las marcas en el DataFrame.
    """
    df[marca_col] = df[marca_col].astype(str).str.strip().str.upper()
    normalized_brands = df[marca_col].apply(normalize_brand)
    
    # Verificar marcas con menos de dos caracteres
    short_brands = normalized_brands[normalized_brands.str.len() < 2].unique()
    for short_brand in short_brands:
        if short_brand:
            response = simpledialog.askstring(
                "Validación de Marca",
                f"La marca '{short_brand}' tiene menos de dos caracteres. ¿Es una marca aparte o un error de digitación? (Escribe 'aparte' o el nombre de la marca correcta)"
            )
            if response and response.lower() != 'aparte':
                response = normalize_brand_name(response)
                if response in normalized_brands.values:
                    normalized_brands = normalized_brands.replace(short_brand, response)
                else:
                    confirm = messagebox.askyesno(
                        "Confirmación de Marca",
                        f"La marca '{response}' no existe en el archivo. ¿Estás seguro de que es el nombre correcto? Si seleccionas 'No', se creará como 'APARTE'."
                    )
                    if confirm:
                        normalized_brands = normalized_brands.replace(short_brand, response)
                    else:
                        normalized_brands = normalized_brands.replace(short_brand, 'APARTE')
    
    df[marca_col] = normalized_brands
    return df

def add_annotation_column(df):
    """
    Agrega una columna de anotaciones al DataFrame solo para el archivo de resultados.
    """
    def check_brand_records(group):
        # Verifica si algún registro de la marca tiene valores nulos o ceros
        relevant_columns = ['CTNS', 'T/CBM', 'T/WEIGHT (KG)']
        has_missing = False
        for _, row in group.iterrows():
            for col in relevant_columns:
                if pd.isnull(row[col]) or row[col] == 0:
                    has_missing = True
                    break
            if has_missing:
                break
        return 'hay registros sin informacion' if has_missing else ''

    # Agregar anotaciones por marca
    marca_col = [col for col in df.columns if 'MARCA' in col.upper()][0]
    annotations = df.groupby(marca_col).apply(check_brand_records)
    df['ANOTACION'] = df[marca_col].map(annotations)
    return df

def process_excel(input_path, output_path, consolidado):
    temp_dir = tempfile.mkdtemp()
    try:
        workbook = load_workbook(input_path)
        df_temp = pd.read_excel(input_path, header=None)
        
        # Buscar ZAFIRO number
        zafiro_number = None
        for idx, row in df_temp.iterrows():
            row_text = ' '.join(str(val) for val in row if pd.notna(val))
            if 'ZAFIRO-' in row_text:
                zafiro_number = re.search(r'(ZAFIRO-\d+-\d+)', row_text).group(1)
                break
        
        header_row = find_header_row(df_temp)
        if header_row is None:
            raise ValueError("No se encontró la fila de encabezados")
        
        images_info = extract_and_save_images_from_workbook(workbook, temp_dir, header_row)
        
        # Leer el DataFrame con los encabezados correctos
        df = pd.read_excel(input_path, header=header_row)
        
        # Limpiar y normalizar nombres de columnas
        df.columns = df.columns.str.strip().str.upper()
        
        # Eliminar columnas Unnamed y normalizar nombres
        normalized_columns = []
        seen_columns = set()
        
        for col in df.columns:
            if 'UNNAMED' in col:
                continue
            clean_col = (col.replace('(', ' ')
                          .replace(')', ' ')
                          .replace('&', ' ')
                          .strip())
            if clean_col in seen_columns:
                continue
            seen_columns.add(clean_col)
            normalized_columns.append(col)
        
        # Mantener solo las columnas normalizadas
        df = df[normalized_columns]
        
        # Identificar columnas de precios para eliminar
        price_columns = [col for col in df.columns 
                        if any(term in col.upper() 
                              for term in ['UNIT PRICE', 'AMOUNT', '单价', '总金额', 'RMB'])]
        df = df.drop(columns=price_columns, errors='ignore')
        
        # Limpiar el DataFrame eliminando filas después de "TOTAL"
        df = clean_dataframe(df)
        
        marca_col = find_brand_column(df)
        if marca_col is None:
            raise ValueError("No se encontró la columna de marca del producto")
        
        # Validar y normalizar marcas
        df = validate_and_normalize_brands(df, marca_col)
        
        # Agregar columna de anotaciones
        df = add_annotation_column(df)
        
        # Verificar si hay registros sin información
        has_missing_info = df['ANOTACION'].str.contains('hay registros sin informacion').any()
        
        # Create results files
        year = str(datetime.now().year)[-2:]
        results_basename = f"RESULTADOS_CONSO_{consolidado}_{year}"
        results_excel = os.path.join(output_path, f"{results_basename}.xlsx")
        results_pdf = os.path.join(output_path, f"{results_basename}.pdf")
        
        # En lugar de copiar directamente, guardamos el DataFrame limpio
        with pd.ExcelWriter(results_excel, engine='openpyxl') as writer:
            # Agregar la hoja de resultados como principal
            create_results_sheet(df, writer, marca_col, has_missing_info)
            
            # Obtener la hoja de resultados y renombrarla a 'Principal'
            worksheet = writer.sheets['RESULTADOS']
            worksheet.title = 'Principal'
            
            # Aplicar formato a las columnas
            header_fill = PatternFill(start_color='ADD8E6', end_color='ADD8E6', fill_type='solid')
            header_font = Font(bold=True)
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
            
            # Aplicar formato a las columnas
            header_fill = PatternFill(start_color='ADD8E6', end_color='ADD8E6', fill_type='solid')
            header_font = Font(bold=True)
            for cell in worksheet[6]:
                cell.fill = header_fill
                cell.font = header_font
            
            # Agregar la hoja de resultados
            create_results_sheet(df, writer, marca_col, has_missing_info)
        
        # Crear PDF solo con la tabla de resultados
        create_pdf_results(results_excel, results_pdf, has_missing_info)
        
        # Procesar archivos por marca
        processed_brands = set()
        
        for marca in df[marca_col].unique():
            if pd.notna(marca) and marca.strip() and marca not in processed_brands:
                processed_brands.add(marca)
                brand_df = df[df[marca_col] == marca].copy()
                try:
                    process_brand_excel(brand_df, output_path, marca, year,
                                     consolidado, images_info, header_row + 1,
                                     zafiro_number)
                    logging.info(f"Archivo creado para marca: {marca}")
                except Exception as e:
                    logging.error(f"Error procesando marca {marca}: {str(e)}")
                    messagebox.showwarning(
                        "Advertencia",
                        f"Hubo un error al procesar la marca {marca}. Por favor, verifique el archivo de salida."
                    )
        
        # Mostrar resumen de procesamiento
        messagebox.showinfo(
            "Proceso Completado",
            f"Se han procesado:\n"
            f"- Archivo de resultados general\n"
            f"- PDF de resultados\n"
            f"- {len(processed_brands)} archivos de marca"
        )
        
    except Exception as e:
        logging.error(f"Error en process_excel: {str(e)}")
        raise
    finally:
        shutil.rmtree(temp_dir)

def create_results_sheet(df, writer, marca_col, has_missing_info):
    """
    Crea la hoja de RESULTADOS en el archivo Excel
    """
    new_column_names = ['SHIPPING MARK MARCA', 'CARTONES', 'CUBICAJE', 'PESO', 'ANOTACION']
    
    # Create summary by brand
    summary = df.groupby(marca_col).agg({
        'CTNS': 'sum',
        'T/CBM': 'sum',
        'T/WEIGHT (KG)': 'sum',
        'ANOTACION': 'first'  # Toma la anotación de la marca
    }).reset_index()
    
    # Remove null annotations
    summary.loc[summary['ANOTACION'].isna(), 'ANOTACION'] = ''
    
    # Write to Excel
    summary.to_excel(writer, sheet_name='RESULTADOS', index=False)
    
    # Get the worksheet to apply formatting
    workbook = writer.book
    worksheet = workbook['RESULTADOS']
    
    # Format headers and change column names
    for col_idx, new_col_name in enumerate(new_column_names, 1):
        cell = worksheet.cell(row=1, column=col_idx)
        cell.value = new_col_name
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    
    # Add totals row
    total_row = len(summary) + 2
    worksheet.cell(row=total_row, column=1, value='TOTAL').font = Font(bold=True)
    
    for col_idx, col_name in enumerate(new_column_names[1:-1], 2):  # Excluir la columna ANOTACION
        col_letter = get_column_letter(col_idx)
        cell = worksheet.cell(row=total_row, column=col_idx)
        cell.value = f'=SUM({col_letter}2:{col_letter}{total_row-1})'
        cell.font = Font(bold=True)
def find_brand_column(df):
    """
    Encuentra la columna que contiene la marca del producto
    """
    possible_names = ['SHIPPING MARK MARCA', 'MARCA DEL PRODUCTO', 'MARCA']
    for col in df.columns:
        col_upper = str(col).upper()
        if any(name in col_upper for name in possible_names):
            return col
    return None

def create_pdf_results(excel_path, pdf_path, has_missing_info):
    try:
        workbook = load_workbook(excel_path, data_only=True)
        sheet = workbook['Principal']
        
        pdf = FPDF()
        pdf.add_page()
        pdf.set_font('Arial', size=10)
        
        # Definimos anchos de columna más pequeños para centrar
        col_widths = [45, 25, 30, 30, 45]  # Reducimos los anchos
        row_height = 8
        page_width = sum(col_widths)
        
        # Calculamos el margen izquierdo para centrar la tabla
        margin_left = (210 - page_width) / 2  # 210 es el ancho de una página A4
        pdf.set_left_margin(margin_left)
        
        pdf.set_font('Arial', 'B', 12)
        pdf.cell(page_width, 10, 'Resultados', 0, 1, 'C')
        pdf.ln(10)
        
        pdf.set_font('Arial', 'B', 10)
        headers = ['SHIPPING MARK MARCA', 'CARTONES', 'CUBICAJE', 'PESO', 'ANOTACION']
        for i, header in enumerate(headers):
            pdf.cell(col_widths[i], row_height, header, 1, 0, 'C')
        pdf.ln(row_height)

        pdf.set_font('Arial', '', 10)
        total_values = [0, 0, 0]

        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0] == 'TOTAL':
                continue

            for idx, value in enumerate(row[1:4]):
                if value not in (None, 'None'):
                    try:
                        total_values[idx] += float(str(value).replace(',', '.'))
                    except (ValueError, TypeError):
                        pass

            for i, value in enumerate(row):
                text = cleanup_text_for_pdf(str(value)) if value not in (None, 'None') else ''
                
                # Reducir el tamaño de la letra para valores con muchos decimales
                if i in [2, 3]:  # Columnas de CUBICAJE y PESO
                    if len(text) > 8:  # Si el número es largo
                        pdf.set_font('Arial', '', 8)  # Reducimos a tamaño 8
                    else:
                        pdf.set_font('Arial', '', 10)  # Volvemos al tamaño normal
                
                pdf.cell(col_widths[i], row_height, text, 1, 0, 'C')
            pdf.ln(row_height)

        # Format totals with smaller font for long numbers
        pdf.set_font('Arial', 'B', 10)
        pdf.cell(col_widths[0], row_height, 'TOTAL', 1, 0, 'C')
        
        # CARTONES total
        total_cartones = f"{total_values[0]:.0f}"
        pdf.cell(col_widths[1], row_height, total_cartones, 1, 0, 'C')
        
        # CUBICAJE total
        total_cubicaje = f"{total_values[1]:.2f}".replace('.', ',')
        font_size = 8 if len(total_cubicaje) > 8 else 10
        pdf.set_font('Arial', 'B', font_size)
        pdf.cell(col_widths[2], row_height, total_cubicaje, 1, 0, 'C')
        
        # PESO total
        total_peso = f"{total_values[2]:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
        font_size = 8 if len(total_peso) > 8 else 10
        pdf.set_font('Arial', 'B', font_size)
        pdf.cell(col_widths[3], row_height, total_peso, 1, 0, 'C')
        
        # Última columna vacía
        pdf.set_font('Arial', 'B', 10)
        pdf.cell(col_widths[4], row_height, '', 1, 0, 'C')
        
        # Añadimos el texto de registros sin información con fuente más pequeña
        if has_missing_info:
            pdf.ln(15)
            pdf.set_left_margin(10)  # Reseteamos el margen
            pdf.set_font('Arial', 'B', 8)  # Reducimos el tamaño de la fuente
            pdf.cell(0, row_height, 'hay registros sin informacion', 0, 1, 'C')
        
        pdf.output(pdf_path)
        
    except Exception as e:
        logging.error(f"Error creating PDF: {str(e)}")
        raise
    
def cleanup_text_for_pdf(text):
    """
    Helper function to clean up text for PDF creation.
    Removes or replaces problematic characters.
    """
    if text is None:
        return ''
    return ''.join(char if ord(char) < 256 else '?' for char in text)
def create_results_file(df, output_path, marca_col):
    results_columns = ['SHIPPING MARK MARCA', 'CTNS', 'T/CBM', 'T/WEIGHT (KG)']
    
    # Create summary by brand
    summary = df.groupby(marca_col).agg({
        'CTNS': 'sum',
        'T/CBM': 'sum',
        'T/WEIGHT (KG)': 'sum'
    }).reset_index()
    
    # Create workbook with results
    wb = Workbook()
    ws = wb.active
    ws.title = 'RESULTADOS'
    
    # Write headers
    for col_idx, col_name in enumerate(results_columns, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = col_name
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    
    # Write data
    for row_idx, row in summary.iterrows():
        for col_idx, col_name in enumerate(results_columns, 1):
            cell = ws.cell(row=row_idx + 2, column=col_idx)
            cell.value = row[col_name if col_name in summary.columns else marca_col]
            cell.alignment = Alignment(horizontal='center')
    
    # Add totals row
    total_row = len(summary) + 2
    ws.cell(row=total_row, column=1).value = 'TOTAL'
    ws.cell(row=total_row, column=1).font = Font(bold=True)
    
    for col_idx, col_name in enumerate(results_columns[1:], 2):
        col_letter = get_column_letter(col_idx)
        ws.cell(row=total_row, column=col_idx).value = f'=SUM({col_letter}2:{col_letter}{total_row-1})'
        ws.cell(row=total_row, column=col_idx).font = Font(bold=True)
    
    wb.save(output_path)

def resize_image(image_path, max_width, max_height):
    """
    Redimensiona la imagen para que se ajuste a las dimensiones máximas permitidas
    sin distorsionarla.
    """
    with PILImage.open(image_path) as img:
        # Calcular proporciones y redimensionar
        img.thumbnail((max_width, max_height), PILImage.LANCZOS)
        temp_path = image_path.replace(".png", "_resized.png")
        img.save(temp_path, "PNG")
        return temp_path

def find_real_header_row(df):
    keywords = ['CTNS', 'MARCA', 'CBM', 'WEIGHT', 'PRODUCTO', 'PRODUCT PICTURE']
    for idx, row in df.iterrows():
        row_str = ' '.join(str(val).upper().strip() for val in row)
        matches = sum(keyword in row_str for keyword in keywords)
        if matches >= 3:  # Aumentamos el número mínimo de coincidencias
            return idx
    return None

def copy_workbook_structure(input_path, header_end_row):
    from openpyxl import load_workbook

    wb = load_workbook(input_path)
    ws = wb.active

    images_info = []
    for image in ws._images:
        images_info.append({
            'image': image,
            'row': image.anchor._from.row + 1,
            'col': image.anchor._from.col + 1
        })

    return wb, images_info

def clean_and_validate_brands(df, marca_col):
    # Normalizar datos (strip espacios y manejo de NaN)
    df[marca_col] = df[marca_col].str.strip()
    
    # Eliminar registros con marcas no válidas
    df = df[df[marca_col].notna()]
    
    # Eliminar caracteres especiales innecesarios
    df[marca_col] = df[marca_col].replace(r'[^\w\s]', '', regex=True)
    
    return df
def extract_and_save_images(df, image_col, output_folder, marca):
    image_folder = os.path.join(output_folder, f"{marca}_imagenes")
    os.makedirs(image_folder, exist_ok=True)

    for i, row in df.iterrows():
        # Obtener el nombre de la imagen
        image_data = row[image_col]
        
        # Verificar si hay datos de imagen en la columna
        if isinstance(image_data, bytes):
            try:
                # Cargar imagen desde los bytes
                image = PILImage.open(BytesIO(image_data))
                
                # Crear nombre de archivo para la imagen
                image_name = f"{marca}_{i+1}.png"
                image_path = os.path.join(output_folder, image_name)
                
                # Guardar la imagen
                image.save(image_path)
                print(f"Imagen guardada: {image_path}")
            except Exception as e:
                print(f"Error al procesar la imagen de la fila {i+1}: {e}")

def main():
    root = tk.Tk()
    app = ExcelProcessorApp(root)
    root.mainloop()

if __name__ == "__main__":
    root = tk.Tk()
    root.withdraw()  # Ocultar la ventana principal de Tkinter
    app = ExcelProcessorApp(root)
    root.mainloop()